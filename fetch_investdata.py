"""
fetch_investdata.py — Fetch mutual fund and index data from Investdata API
==========================================================================
Downloads D-1, D-2, and INDEKS data from the Infovesta API and writes
Excel files matching the format expected by raw_to_compiled.py.

Reduces manual file input from 6 to 3:
  MANUAL (keep): HistoricalNAV_280426.xlsx, HistoricalNAV_270426_2.xlsx, indeks_bloomberg.xlsx
  API (auto):    Download Data D-1, Download Data D-2, Indeks

Usage:
    python fetch_investdata.py [output_dir]

Output (in very_raw_data/ or specified dir):
  Download Data Reksadana Per <date>.xlsx  — D-1 sheet format
  Download Data Reksadana Per <date2>.xlsx — D-2 sheet format
  Indeks_<date>.xlsx                        — INDEKS sheet format

Credentials from investdata_api/.env
"""

import os
import sys
import json
import time
import requests
import pandas as pd
import openpyxl
from datetime import datetime

# ─── Config ──────────────────────────────────────────────────────────────────
_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ENV_FILE = os.path.join(_BASE_DIR, "investdata_api", ".env")
TOKEN_FILE = os.path.join(_BASE_DIR, "investdata_api", "token.json")
OAUTH_URL = "https://api.infovesta.com/gateway/oauth/token"
API_BASE = "https://api.infovesta.com"
OUTPUT_DIR = os.path.join(_BASE_DIR, "very_raw_data")


def load_env():
    """Load credentials from environment variables or .env file fallback."""
    # Prefer environment variables (Render deployment)
    if os.environ.get("INVESTDATA_USERNAME"):
        return {
            "INVESTDATA_USERNAME": os.environ["INVESTDATA_USERNAME"],
            "INVESTDATA_PASSWORD": os.environ["INVESTDATA_PASSWORD"],
            "INVESTDATA_CLIENT_ID": os.environ.get("INVESTDATA_CLIENT_ID", "api2"),
            "INVESTDATA_CLIENT_SECRET": os.environ.get("INVESTDATA_CLIENT_SECRET", "api2"),
        }
    # Fallback to .env file (local development)
    env = {}
    with open(ENV_FILE) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            key, _, val = line.partition("=")
            env[key.strip()] = val.strip()
    return env


def get_token():
    """Get a valid access token (refresh if expired)."""
    env = load_env()

    # Use /tmp for token on Vercel (read-only filesystem at /var/task)
    token_file = TOKEN_FILE
    if os.environ.get("VERCEL"):
        token_file = "/tmp/investdata_token.json"

    # Try existing token first
    if os.path.exists(token_file):
        with open(token_file) as f:
            tok = json.load(f)
        # expires_in is a static number — check actual age using obtained_at
        obtained_at = tok.get("obtained_at", 0)
        expires_in = tok.get("expires_in", 0)
        if obtained_at > 0 and (time.time() - obtained_at) < (expires_in - 300):
            return tok["access_token"]

    # Get fresh token
    resp = requests.post(
        OAUTH_URL,
        auth=(env["INVESTDATA_CLIENT_ID"], env["INVESTDATA_CLIENT_SECRET"]),
        data={
            "grant_type": "password",
            "username": env["INVESTDATA_USERNAME"],
            "password": env["INVESTDATA_PASSWORD"],
            "scope": "read",
        },
    )
    resp.raise_for_status()
    tok = resp.json()
    tok["obtained_at"] = time.time()
    # Persist token (skip on Vercel if /tmp not writable)
    try:
        os.makedirs(os.path.dirname(token_file), exist_ok=True)
        with open(token_file, "w") as f:
            json.dump(tok, f, indent=2)
    except OSError:
        pass
    return tok["access_token"]


def fetch_all(token):
    """Fetch all required data from the API."""
    headers = {"Authorization": f"Bearer {token}"}

    # 1. Daily NAV (latest 2 dates — gives both D-1 and D-2 data)
    r = requests.get(f"{API_BASE}/api/mutualfund/data/dailynavlatesttwodates", headers=headers)
    r.raise_for_status()
    nav_data = r.json()

    # 2. AUM data
    r = requests.get(f"{API_BASE}/api/mutualfund/data/aum", headers=headers)
    r.raise_for_status()
    aum_data = r.json()

    # 3. Scoring data
    r = requests.get(f"{API_BASE}/api/mutualfund/data/scoring", headers=headers)
    r.raise_for_status()
    scoring_data = r.json()

    # 4. Indeks daily price
    r = requests.get(f"{API_BASE}/api/indeks/data/dailyprice", headers=headers)
    r.raise_for_status()
    indeks_data = r.json()

    return nav_data, aum_data, scoring_data, indeks_data


def build_download_sheet(nav_data, aum_data, scoring_data, target_date_str):
    """Build a Download Data-style DataFrame for a specific date.

    target_date_str: 'YYYY-MM-DD' — pick the most recent (D-1) or second (D-2)
    """
    # Filter nav data to target date
    nav_by_date = [r for r in nav_data if r["date"] == target_date_str]
    print(f"  NAV records for {target_date_str}: {len(nav_by_date)}")

    # Find the scoring date (from scoring data, closest to target)
    scoring_dates = sorted(set(r["date"] for r in scoring_data), reverse=True)
    scoring_date = scoring_dates[0] if scoring_dates else target_date_str
    # Format as DD-Mon-YYYY for column names
    dt = datetime.strptime(scoring_date, "%Y-%m-%d")
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    score_label = f"{dt.day}-{months[dt.month - 1]}-{dt.year}"

    # Index AUM by productId
    aum_by_product = {}
    for r in aum_data:
        aum_by_product[r["productId"]] = r

    # Index scoring by productId
    scoring_by_product = {}
    for r in scoring_data:
        scoring_by_product[r["productId"]] = r

    rows = []
    for nav in nav_by_date:
        pid = nav["productId"]
        aum = aum_by_product.get(pid, {})
        score = scoring_by_product.get(pid, {})

        # Get AUM date from the aum record
        aum_date_label = score_label
        if aum.get("date"):
            a_dt = datetime.strptime(aum["date"], "%Y-%m-%d")
            aum_date_label = f"{a_dt.day}-{months[a_dt.month - 1]}-{a_dt.year}"

        row = {
            "Nama": nav["name"],
            "Jenis": nav["typeName"],
            " Kode Manajer Investasi": nav.get("fundmgrId", ""),
            "Manajer Investasi": nav.get("fundmgr", ""),
            "Kustodian": nav.get("custodian", ""),
            "Denominasi": nav.get("currency", ""),
            "Deviden": "",
            "Syariah": "Syariah" if nav.get("syariah") else "Non Syariah",
            "NAB/UP": nav.get("nav"),
            "1 Hr(%)": nav.get("onedayreturn"),
            "1 Mgg(%)": nav.get("oneweekreturn"),
            "MTD(%)": nav.get("mtdreturn"),
            "1 Bln(%)": nav.get("onemonthreturn"),
            "3 Bln(%)": nav.get("threemonthreturn"),
            "6 Bln(%)": nav.get("sixmonthreturn"),
            "YTD(%)": nav.get("ytdreturn"),
            "1 Thn(%)": nav.get("oneyearreturn"),
            "3 Thn(%)": nav.get("threeyearreturn"),
            "5 Thn(%)": nav.get("fiveyearreturn"),
            "10 Thn(%)": None,
            f"Scoring 6 Bulan ({score_label})": score.get("ranksixmonth"),
            f"Scoring 1 Tahun ({score_label})": score.get("rankoneyear"),
            f"Scoring 3 Tahun ({score_label})": score.get("rankthreeyear"),
            f"Scoring 5 Tahun ({score_label})": score.get("rankfiveyear"),
            f"AUM ({aum_date_label})": aum.get("aum"),
            f"Unit ({aum_date_label})": aum.get("unit"),
            f"AUM MI({aum_date_label})": None,
            f"AUM MI dengan Penyertaan Terbatas({aum_date_label})": None,
            f"AUM MI Dollar({aum_date_label})": None,
            f"AUM MI Dollar dengan Penyertaan Terbatas({aum_date_label})": None,
            "Kepemilikan": "",
            "Tanggal Penawaran Umum": None,
            "Kode ISIN": nav.get("isinCode", ""),
            "Total AUM Shared Class": None,
        }
        rows.append(row)

    df = pd.DataFrame(rows)

    # Copy AUM from base names to their TR variants (API only provides AUM for base names)
    aum_col_key = f"AUM ({aum_date_label})"
    unit_col_key = f"Unit ({aum_date_label})"
    import re
    for i, row in df.iterrows():
        name = str(row.get("Nama", ""))
        if re.search(r" - total return\*?$", name, re.IGNORECASE):
            base_name = re.sub(r"\s*-\s*total return\*?\s*$", "", name, flags=re.IGNORECASE).strip()
            base_row = df[df["Nama"] == base_name]
            if len(base_row) > 0 and pd.isna(row.get(aum_col_key)):
                df.at[i, aum_col_key] = base_row.iloc[0].get(aum_col_key)
                df.at[i, unit_col_key] = base_row.iloc[0].get(unit_col_key)

    return df


def build_indeks_sheet(indeks_data):
    """Build an INDEKS-style DataFrame."""
    rows = []
    for item in indeks_data:
        rows.append({
            "Nama": item.get("productName") or item.get("productId", ""),
            "Nilai": item.get("price"),
            # API returns percentage (e.g. -2.0086), original uses decimal (e.g. -0.020086)
            "1 Hr": item.get("onedayreturn") / 100 if item.get("onedayreturn") is not None else None,
            "1 Mgg": item.get("oneweekreturn") / 100 if item.get("oneweekreturn") is not None else None,
            "MTD": item.get("mtdreturn") / 100 if item.get("mtdreturn") is not None else None,
            "1 Bln": item.get("onemonthreturn") / 100 if item.get("onemonthreturn") is not None else None,
            "3 Bln": item.get("threemonthreturn") / 100 if item.get("threemonthreturn") is not None else None,
            "6 Bln": item.get("sixmonthreturn") / 100 if item.get("sixmonthreturn") is not None else None,
            "YTD": item.get("ytdreturn") / 100 if item.get("ytdreturn") is not None else None,
            "1 Thn": item.get("oneyearreturn") / 100 if item.get("oneyearreturn") is not None else None,
            "3 Thn": item.get("threeyearreturn") / 100 if item.get("threeyearreturn") is not None else None,
            "5 Thn": item.get("fiveyearreturn") / 100 if item.get("fiveyearreturn") is not None else None,
        })
    return pd.DataFrame(rows)


def date_label(date_str):
    """Convert 'YYYY-MM-DD' to 'DD-Mon-YYYY' format for filenames."""
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    return f"{dt.day}-{months[dt.month - 1]}-{dt.year}"


def write_download_excel(df, path, title):
    """Write Download Data Excel with title row at row 1, headers at row 2."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "sheet 0"
    ws.cell(1, 1, title)
    for ci, col in enumerate(df.columns, 1):
        ws.cell(2, ci, col)
    for ri, row in enumerate(df.values, 3):
        for ci, val in enumerate(row, 1):
            ws.cell(ri, ci, val)
    wb.save(path)


def write_indeks_excel(df, path, title):
    """Write INDEKS Excel with title row at row 1, headers at row 2."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "sheet 0"
    ws.cell(1, 1, title)
    for ci, col in enumerate(df.columns, 1):
        ws.cell(2, ci, col)
    for ri, row in enumerate(df.values, 3):
        for ci, val in enumerate(row, 1):
            ws.cell(ri, ci, val)
    wb.save(path)


def main():
    if len(sys.argv) > 1:
        out_dir = sys.argv[1]
    else:
        out_dir = OUTPUT_DIR

    os.makedirs(out_dir, exist_ok=True)

    print("Fetching data from Investdata API...")
    token = get_token()
    nav_data, aum_data, scoring_data, indeks_data = fetch_all(token)
    print(f"Total NAV records: {len(nav_data)}")
    print(f"Total AUM records: {len(aum_data)}")
    print(f"Total Scoring records: {len(scoring_data)}")
    print(f"Total Indeks records: {len(indeks_data)}")

    # Determine the two most recent dates from NAV data
    dates = sorted(set(r["date"] for r in nav_data), reverse=True)
    if len(dates) < 2:
        print(f"ERROR: Need at least 2 dates, got {dates}")
        sys.exit(1)
    d1_date = dates[0]
    d2_date = dates[1]
    print(f"D-1 date: {d1_date}")
    print(f"D-2 date: {d2_date}")

    # Build D-1 sheet
    print(f"Building D-1 for {d1_date}...")
    d1_df = build_download_sheet(nav_data, aum_data, scoring_data, d1_date)
    d1_label = date_label(d1_date)
    d1_path = os.path.join(out_dir, f"Download Data Reksadana Per {d1_label}.xlsx")
    write_download_excel(d1_df, d1_path, f"Data Per Tanggal : {d1_date}")
    print(f"  Wrote {d1_path} ({len(d1_df)} funds)")

    # Build D-2 sheet
    print(f"Building D-2 for {d2_date}...")
    d2_df = build_download_sheet(nav_data, aum_data, scoring_data, d2_date)
    d2_label = date_label(d2_date)
    d2_path = os.path.join(out_dir, f"Download Data Reksadana Per {d2_label}.xlsx")
    write_download_excel(d2_df, d2_path, f"Data Per Tanggal : {d2_date}")
    print(f"  Wrote {d2_path} ({len(d2_df)} funds)")

    # Build INDEKS sheet
    print("Building INDEKS...")
    indeks_df = build_indeks_sheet(indeks_data)
    indeks_label = date_label(d1_date)
    indeks_path = os.path.join(out_dir, f"Indeks_{indeks_label}.xlsx")
    write_indeks_excel(indeks_df, indeks_path, f"Data Per {d1_date}")
    print(f"  Wrote {indeks_path} ({len(indeks_df)} entries)")

    # Clean up old files for previous dates
    for f in os.listdir(out_dir):
        if f.startswith("Download Data Reksadana Per ") and f.endswith(".xlsx"):
            if d1_label not in f and d2_label not in f:
                old_path = os.path.join(out_dir, f)
                os.remove(old_path)
                print(f"  Removed old: {f}")
        if f.startswith("Indeks_") and f.endswith(".xlsx"):
            if indeks_label not in f:
                old_path = os.path.join(out_dir, f)
                os.remove(old_path)
                print(f"  Removed old: {f}")

    print("\nDone. Required manual files:")
    print(f"  {out_dir}/HistoricalNAV_*.xlsx (from BRI)")
    print(f"  {out_dir}/indeks_bloomberg.xlsx (from Bloomberg)")


if __name__ == "__main__":
    main()
