"""
Parity Check: Compare brimi_output.xlsx vs original Excel (NEW PAGE TABLE)
==========================================================================
Compares row-by-row in document order. Since the same fund can appear in
multiple sections, we compare by positional order, not by name.

Usage:
    python parity_check.py <original_excel> <output_excel>
"""
import sys, math, re
import openpyxl


def fmt(v):
    """Normalize a cell value for comparison."""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v.lower() in ("n/a", "#n/a", "#ref!", "#value!", "", "none", "nan"):
            return None
        try:
            return round(float(v), 4)
        except ValueError:
            return v
    if isinstance(v, (int, float)):
        if math.isnan(v) or math.isinf(v):
            return None
        return round(float(v), 4)
    return v


def close_enough(a, b, tol=0.02):
    """Are two numeric values within tolerance?"""
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    if isinstance(a, str) or isinstance(b, str):
        return str(a) == str(b)
    if a == 0 and b == 0:
        return True
    denom = max(abs(a), abs(b), 1e-10)
    return abs(a - b) / denom < tol


def extract_fund_rows(ws):
    """
    Extract fund data rows from a page table sheet, preserving order.
    Returns list of dicts with keys:
      name, nav, 1hr, 1mgg, mtd, 1bln, 3bln, 6bln, ytd, 1thn, 3thn, rank, aum
    Skips section headers, group headers, column headers — only data rows.
    """
    funds = []
    for row in ws.iter_rows(min_row=1, max_row=500, values_only=True):
        if not row or len(row) < 13:
            continue
        a = row[0]
        b = row[1]
        if a is None or b is None:
            continue
        if isinstance(a, str) and a.strip() in ("Nama", ""):
            continue
        if not isinstance(b, (int, float)):
            continue
        # This is a data row
        funds.append({
            "name":   str(a).strip(),
            "nav":    fmt(b),
            "1hr":    fmt(row[2]) if len(row) > 2 else None,
            "1mgg":   fmt(row[3]) if len(row) > 3 else None,
            "mtd":    fmt(row[4]) if len(row) > 4 else None,
            "1bln":   fmt(row[5]) if len(row) > 5 else None,
            "3bln":   fmt(row[6]) if len(row) > 6 else None,
            "6bln":   fmt(row[7]) if len(row) > 7 else None,
            "ytd":    fmt(row[8]) if len(row) > 8 else None,
            "1thn":   fmt(row[9]) if len(row) > 9 else None,
            "3thn":   fmt(row[10]) if len(row) > 10 else None,
            "rank":   fmt(row[11]) if len(row) > 11 else None,
            "aum":    fmt(row[12]) if len(row) > 12 else None,
        })
    return funds


FIELDS = ["nav", "1hr", "1mgg", "mtd", "1bln", "3bln", "6bln", "ytd", "1thn", "3thn", "rank", "aum"]
FIELD_LABELS = {
    "nav": "NAV", "1hr": "1Hr", "1mgg": "1Mgg", "mtd": "MTD", "1bln": "1Bln",
    "3bln": "3Bln", "6bln": "6Bln", "ytd": "YTD", "1thn": "1Thn", "3thn": "3Thn",
    "rank": "Rank", "aum": "AUM"
}


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    orig_path = sys.argv[1]
    out_path  = sys.argv[2]

    print(f"Loading original : {orig_path}")
    wb_orig = openpyxl.load_workbook(orig_path, data_only=True)
    print(f"Loading output   : {out_path}")
    wb_out  = openpyxl.load_workbook(out_path, data_only=True)

    # --- PAGE TABLE ---
    ws_orig = wb_orig["NEW PAGE TABLE"]
    ws_out  = wb_out["NEW PAGE TABLE (output)"]

    orig_funds = extract_fund_rows(ws_orig)
    out_funds  = extract_fund_rows(ws_out)

    print(f"\nOriginal fund rows : {len(orig_funds)}")
    print(f"Output fund rows   : {len(out_funds)}")

    # --- Row count comparison ---
    if len(orig_funds) != len(out_funds):
        print(f"\n⚠ ROW COUNT MISMATCH: original has {len(orig_funds)}, output has {len(out_funds)}")
        # Find missing/extra by comparing name sequences
        orig_names = [f["name"] for f in orig_funds]
        out_names  = [f["name"] for f in out_funds]
        orig_set = set(orig_names)
        out_set  = set(out_names)
        missing = [n for n in orig_names if n not in out_set]
        extra   = [n for n in out_names if n not in orig_set]
        if missing:
            print(f"  Missing from output ({len(missing)}):")
            for n in sorted(set(missing)):
                count = missing.count(n)
                print(f"    {n}" + (f" (×{count})" if count > 1 else ""))
        if extra:
            print(f"  Extra in output ({len(extra)}):")
            for n in sorted(set(extra)):
                count = extra.count(n)
                print(f"    {n}" + (f" (×{count})" if count > 1 else ""))

    # --- Positional comparison ---
    # Compare row-by-row up to the minimum length
    total_cells = 0
    matched = 0
    mismatched = 0
    mismatches = []

    compare_len = min(len(orig_funds), len(out_funds))
    for pos in range(compare_len):
        orig_f = orig_funds[pos]
        out_f  = out_funds[pos]

        if orig_f["name"] != out_f["name"]:
            mismatches.append({
                "pos": pos + 1,
                "name": f"ROW MISMATCH: orig='{orig_f['name']}' vs ours='{out_f['name']}'",
                "field": "Name",
                "orig": orig_f["name"],
                "ours": out_f["name"],
            })
            continue

        for field in FIELDS:
            total_cells += 1
            orig_val = orig_f[field]
            out_val  = out_f[field]

            if close_enough(orig_val, out_val):
                matched += 1
            else:
                mismatched += 1
                mismatches.append({
                    "pos": pos + 1,
                    "name": orig_f["name"],
                    "field": FIELD_LABELS[field],
                    "orig": orig_val,
                    "ours": out_val,
                })

    # --- SUMMARY ---
    print(f"\n{'='*80}")
    print(f"PARITY CHECK SUMMARY — NEW PAGE TABLE (positional)")
    print(f"{'='*80}")
    print(f"  Rows compared       : {compare_len}")
    print(f"  Total cells compared : {total_cells}")
    print(f"  Matched              : {matched}")
    print(f"  Mismatched           : {mismatched}")
    pct = (matched / total_cells * 100) if total_cells else 0
    print(f"  Match rate           : {pct:.1f}%")

    if mismatches:
        print(f"\n{'─'*80}")
        print(f"MISMATCHES ({len(mismatches)}):")
        print(f"{'─'*80}")
        print(f"{'#':>4} {'Fund Name':<55} {'Field':<6} {'Original':>12} {'Ours':>12}")
        print(f"{'─'*4} {'─'*55} {'─'*6} {'─'*12} {'─'*12}")
        for m in mismatches:
            ov = f"{m['orig']}" if m['orig'] is not None else "None"
            uv = f"{m['ours']}" if m['ours'] is not None else "None"
            print(f"{m['pos']:4d} {m['name'][:55]:<55} {m['field']:<6} {ov:>12} {uv:>12}")

    # Count mismatches by type
    if mismatches:
        print(f"\n{'─'*80}")
        print("MISMATCH BREAKDOWN BY FIELD:")
        print(f"{'─'*80}")
        from collections import Counter
        by_field = Counter(m["field"] for m in mismatches)
        for field, count in sorted(by_field.items(), key=lambda x: -x[1]):
            print(f"  {field:<10}: {count} mismatches")

        none_vs_value = sum(1 for m in mismatches
                           if (m["orig"] is None) != (m["ours"] is None))
        value_vs_value = len(mismatches) - none_vs_value
        print(f"\n  None vs Value : {none_vs_value}")
        print(f"  Value vs Value: {value_vs_value}")


if __name__ == "__main__":
    main()
