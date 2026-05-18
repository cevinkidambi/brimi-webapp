"""
BRIMI Fund Performance & AUM Processor
=======================================
Reads raw data sheets and produces:
  • NEW PAGE TABLE (output)        — full ranked fund table by category
  • NEW PERFORMANCE TABLE (output) — BRI-managed funds with Score & Quartile

Usage:
    python process_brimi.py <input_excel> [output_excel] [fund_universe_json] [fund_map_json]

Defaults:
    output_excel       = brimi_output.xlsx
    fund_universe_json = fund_universe.json
    fund_map_json      = fund_map.json  (optional)

To add/remove funds edit fund_universe.json directly, or use:
    python manage_funds.py --help
"""

import sys, re, json, os, warnings, math
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

warnings.filterwarnings("ignore")


# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

PERF_COLS = ["1 Hr(%)", "1 Mgg(%)", "MTD(%)", "1 Bln(%)", "3 Bln(%)",
             "6 Bln(%)", "YTD(%)", "1 Thn(%)", "3 Thn(%)", "5 Thn(%)"]

D1_COL = {
    "NAB/UP":   "NAB/UP",
    "1 Hr(%)":  "1 Hr(%)",
    "1 Mgg(%)": "1 Mgg(%)",
    "MTD(%)":   "MTD(%)",
    "1 Bln(%)": "1 Bln(%)",
    "3 Bln(%)": "3 Bln(%)",
    "6 Bln(%)": "6 Bln(%)",
    "YTD(%)":   "YTD(%)",
    "1 Thn(%)": "1 Thn(%)",
    "3 Thn(%)": "3 Thn(%)",
    "5 Thn(%)": "5 Thn(%)",
    "AUM":      None,   # detected dynamically from D-1 header
}

# Sections whose formulas reference D-2 instead of D-1
D2_SECTION_KEYWORDS = ["global sharia equity"]

INDEKS_COL = {
    "NAB/UP":   "Nilai",
    "1 Hr(%)":  "1 Hr",
    "1 Mgg(%)": "1 Mgg",
    "MTD(%)":   "MTD",
    "1 Bln(%)": "1 Bln",
    "3 Bln(%)": "3 Bln",
    "6 Bln(%)": "6 Bln",
    "YTD(%)":   "YTD",
    "1 Thn(%)": "1 Thn",
    "3 Thn(%)": "3 Thn",
    "5 Thn(%)": None,
}

# Bloomberg sheet: fund name → 0-based row index
BLOOMBERG_ROW = {
    "MSCI World Islamic Index":            8,
    "MSCI Indonesia ESG Screened":         7,
    "INDOBex Government Total Return*":    9,
}
BB_COL = {
    "I": "1 Hr(%)", "J": "1 Mgg(%)", "K": "MTD(%)",  "L": "1 Bln(%)",
    "M": "3 Bln(%)", "N": "6 Bln(%)", "O": "YTD(%)", "P": "1 Thn(%)",
    "Q": "3 Thn(%)",
}

INDEKS_NAMES = {
    "LQ45", "IDX Small-Mid Cap Liquid", "Indeks Harga Saham Gabungan",
    "Indeks Saham Syariah Indonesia", "Infovesta sharia Equity fund Index",
    "Jakarta Islamic Index", "Infovesta Money Market Fund Index",
    "Infovesta Balanced Fund Index", "Infovesta Government Bond Index",
    "Infovesta Government Bond Index Short", "Infovesta Sharia Balanced Fund Index",
    "Infovesta Sharia Money Market Fund", "Infovesta Fixed Income Fund Index",
    "MSCI Indonesia ESG Screened", "Papan Akselerasi",
}

BLOOMBERG_NAMES = {"MSCI World Islamic Index", "MSCI Indonesia ESG Screened",
                   "INDOBex Government Total Return*"}

QUARTILE_LABELS = {1: "Q1", 2: "Q2", 3: "Q3", 4: "Q4"}
QUARTILE_COLORS = {"Q1": "70AD47", "Q2": "FFC000", "Q3": "ED7D31", "Q4": "FF0000"}

FUND_MAP: dict = {}

# Scoring API score string → star text format
SCORE_STARS = {
    "0": "No Score", "1-": "* -", "1": "*", "1+": "* +",
    "2-": "* * -", "2": "* *", "2+": "* * +",
    "3-": "* * * -", "3": "* * *", "3+": "* * * +",
    "4-": "* * * * -", "4": "* * * *", "4+": "* * * * +",
    "5-": "* * * * * -", "5": "* * * * *", "5+": "* * * * * +",
    "notapplied": "No Score",
}


# ─────────────────────────────────────────────────────────────────────────────
# NORMALISATION
# ─────────────────────────────────────────────────────────────────────────────

def norm(s: str) -> str:
    return str(s).strip().lower()

def strip_tr(name: str) -> str:
    """Strip '- Total Return' or '- Price Return' suffix and trailing asterisks."""
    s = re.sub(r"\s*-\s*(?:total|price) return\*?\s*$", "", name, flags=re.IGNORECASE).strip()
    return re.sub(r"\*+\s*$", "", s).strip()


# ─────────────────────────────────────────────────────────────────────────────
# LOOKUP HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def build_lookup(df: pd.DataFrame, name_col: str = "Nama") -> dict:
    """
    {norm(name): row} — each fund name maps to its own row.
    No TR auto-promotion: base names and TR names are kept separate.
    """
    lkp: dict = {}
    for _, row in df.iterrows():
        nm = str(row[name_col]).strip() if pd.notna(row[name_col]) else ""
        if not nm or nm in ("nan", "None"):
            continue
        k = norm(nm)
        if k not in lkp:
            lkp[k] = row
    return lkp


def lookup(name: str, lkp: dict):
    """Try exact → strip TR → None."""
    k = norm(name)
    if k in lkp:
        return lkp[k]
    return lkp.get(norm(strip_tr(name)))


def lookup_notr(name: str, lkp: dict):
    """Explicitly return the non-TR (base) row, ignoring TR-promoted entries."""
    for candidate in (norm(name), norm(strip_tr(name))):
        row = lkp.get(candidate)
        if row is not None:
            row_name = str(row.get("Nama", "")).strip()
            if not re.search(r"total return\*?$", row_name, re.IGNORECASE):
                return row
    return None


def safe(row, col):
    if col is None or row is None:
        return None
    try:
        v = row[col]
        if v is None or (isinstance(v, float) and math.isnan(v)):
            return None
        if isinstance(v, str) and v.strip().upper() in ("N/A", "#N/A", "#REF!", "#VALUE!"):
            return None
        return v
    except (KeyError, TypeError):
        return None


def _fmt(v):
    if v is None:
        return None
    try:
        f = float(v)
        return None if math.isnan(f) else round(f, 4)
    except (TypeError, ValueError):
        return None


# ─────────────────────────────────────────────────────────────────────────────
# CONFIG LOADERS
# ─────────────────────────────────────────────────────────────────────────────

def load_fund_map(path: str) -> dict:
    if not os.path.exists(path):
        return {}
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    return {
        norm(e["display_name"]): {
            "brimi_name": e.get("brimi_name"),
            "d1_name":    e.get("d1_name"),
            "nav_name":   e.get("nav_name"),
        }
        for e in data.get("mappings", [])
    }


def load_universe(path: str) -> dict:
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Fund universe file not found: {path}\n"
            "Generate it with: python manage_funds.py --init <input_excel>"
        )
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def _section_source(section: dict) -> str:
    """Determine whether a section uses D-1 or D-2 as its primary source."""
    name_lc = section["section"].lower()
    if any(kw in name_lc for kw in D2_SECTION_KEYWORDS):
        return "d2"
    return "d1"


def _apply_peer_overrides(universe: dict):
    """Merge peer_overrides.json into universe page_table sections.

    Overrides stored as {"0": [funds], "5": [funds]} — dict keyed by section index.
    Committed via GitHub API from the admin page.
    """
    overrides_path = os.path.join(os.path.dirname(__file__), "peer_overrides.json")
    if not os.path.exists(overrides_path):
        return

    with open(overrides_path) as f:
        overrides = json.load(f)

    sections = universe["page_table"]["sections"]
    for idx_str, funds in overrides.items():
        idx = int(idx_str)
        if 0 <= idx < len(sections):
            sections[idx]["funds"] = funds


def get_page_table_config(universe: dict) -> list[dict]:
    """Flatten universe → page_table into a list of fund dicts.
    Each entry gets a unique section_idx for independent ranking."""
    _apply_peer_overrides(universe)
    entries = []
    for si, section in enumerate(universe["page_table"]["sections"]):
        source = _section_source(section)
        for fund in section["funds"]:
            if not fund.get("active", True):
                continue
            entries.append({
                "section":      section["section"],
                "section_idx":  si,
                "rank_group":   section.get("rank_group"),
                "display_name": fund["display_name"],
                "alias":        fund.get("alias"),
                "is_usd":       section.get("is_usd", False),
                "source":       source,
                "is_index":     fund.get("is_index", False),
                "aum_unit":     section.get("aum_unit", "Rp Miliar"),
            })
    return entries


def get_perf_table_config(universe: dict) -> list[dict]:
    """Return active funds from universe → performance_table."""
    return [f for f in universe["performance_table"]["funds"] if f.get("active", True)]


# ─────────────────────────────────────────────────────────────────────────────
# BLOOMBERG LOADER
# ─────────────────────────────────────────────────────────────────────────────

def load_bloomberg(wb_in) -> dict:
    out = {}
    if "BloombergIndex" not in wb_in.sheetnames:
        return out
    ws   = wb_in["BloombergIndex"]
    hdr  = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    cmap = {get_column_letter(j + 1): j for j in range(len(hdr))}
    rows = list(ws.iter_rows(values_only=True))
    for name, ridx in BLOOMBERG_ROW.items():
        if ridx < len(rows):
            r = rows[ridx]
            entry = {col: r[cmap[letter]] if cmap.get(letter) is not None else None
                           for letter, col in BB_COL.items()}
            h = cmap.get("H")
            entry["NAB/UP"] = r[h] if h is not None and h < len(r) else None
            out[name] = entry
    return out


# ─────────────────────────────────────────────────────────────────────────────
# PERFORMANCE RESOLVER
# ─────────────────────────────────────────────────────────────────────────────

def _resolve_perf_name(display_name: str, alias: str | None) -> str:
    """Determine the lookup name for performance columns.

    Rule (from original formula analysis):
    - If alias is None → use display_name
    - If alias == strip_tr(display_name) → alias is just the base form
      of a TR display name, ignore it for perf → use display_name
    - Otherwise → use alias (it provides a genuinely different lookup target)
    """
    if not alias:
        return display_name
    if norm(alias) == norm(strip_tr(display_name)):
        return display_name
    return alias


def get_perf(name: str, alias: str | None, source: str,
             d1_lkp: dict, d2_lkp: dict, indeks_lkp: dict,
             bloomberg_data: dict, is_price_return: bool = False) -> dict:
    """Resolve NAV + performance columns for a fund.

    Args:
        name:   display_name from the template
        alias:  col N alias from the template (or None)
        source: "d1" or "d2" — which sheet the original formula references
    """
    result = {c: None for c in ["NAB/UP"] + PERF_COLS}

    if name in BLOOMBERG_NAMES:
        result.update(bloomberg_data.get(name, {}))
        return result

    if name in INDEKS_NAMES:
        row = lookup(name, indeks_lkp)
        if row is not None:
            result["NAB/UP"] = safe(row, INDEKS_COL["NAB/UP"])
            for col in PERF_COLS:
                result[col] = safe(row, INDEKS_COL.get(col))
        return result

    override = FUND_MAP.get(norm(name), {})

    # --- NAV: always use display_name (or nav_name override) ---
    nav_name = override.get("nav_name") or name
    primary   = d2_lkp if source == "d2" else d1_lkp
    secondary = d1_lkp if source == "d2" else d2_lkp

    fn = lookup_notr if is_price_return else lookup
    nav_row = fn(nav_name, primary)
    if nav_row is None:
        nav_row = fn(nav_name, secondary)
    if nav_row is not None:
        result["NAB/UP"] = safe(nav_row, D1_COL["NAB/UP"])

    # --- Performance: use alias rule ---
    perf_name = _resolve_perf_name(name, alias)
    perf_row = fn(perf_name, primary)
    if perf_row is None:
        perf_row = fn(perf_name, secondary)
    if perf_row is not None:
        for col in PERF_COLS:
            result[col] = safe(perf_row, D1_COL.get(col))

    return result


# ─────────────────────────────────────────────────────────────────────────────
# AUM RESOLVER
# ─────────────────────────────────────────────────────────────────────────────

def get_aum(name: str, brimi_alias: str | None, is_usd: bool,
            brimi_d1_lkp: dict, brimi_d2_lkp: dict,
            d1_lkp: dict, d2_lkp: dict) -> float | None:
    """
    Priority:
      1. BRIMI D-1 / D-2 — BRI's own daily AUM (primary)
      2. D-1 / D-2 AUM column — our data (fallback)
    """
    override   = FUND_MAP.get(norm(name), {})
    brimi_name = override.get("brimi_name") or brimi_alias or name
    d1_name    = override.get("d1_name")    or brimi_alias or name
    divisor    = 1e6 if is_usd else 1e9

    # 1. BRIMI primary
    brimi_lkp = brimi_d2_lkp if is_usd else brimi_d1_lkp
    for nm in _uniq(brimi_name, name):
        row = lookup(nm, brimi_lkp)
        if row is not None:
            v = safe(row, "AUM")
            if v is not None and float(v) != 0:
                return float(v) / divisor

    # 2. Cross-check BRIMI D-1 for USD funds
    if is_usd:
        for nm in _uniq(brimi_name, name):
            row = lookup(nm, brimi_d1_lkp)
            if row is not None:
                v = safe(row, "AUM")
                if v is not None and float(v) != 0:
                    return float(v) / 1e9

    # 3. Fallback D-1/D-2
    lkp = d2_lkp if is_usd else d1_lkp
    for nm in _uniq(d1_name, name):
        row = lookup(nm, lkp)
        if row is not None:
            v = safe(row, D1_COL["AUM"])
            if v is not None:
                return float(v) / divisor

    return None


def _uniq(*args) -> list:
    seen, out = set(), []
    for a in args:
        if a and a not in seen:
            seen.add(a); out.append(a)
    return out


# ─────────────────────────────────────────────────────────────────────────────
# RANKING & QUARTILE
# ─────────────────────────────────────────────────────────────────────────────

def rank_quartile(values: list) -> list[tuple]:
    """Rank values descending, assign quartile using Excel QUARTILE thresholds on ranks.

    Matches the reference Excel:
      1. RANK(value, range) — ties get same rank (RANK.EQ behavior)
      2. T_q = QUARTILE(ranks, q) — uses PERCENTILE.INC interpolation
      3. Q = IF(rank < T1, 1, IF(rank < T2, 2, IF(rank < T3, 3, 4)))
    """
    n = len(values)
    valid = [(i, float(v)) for i, v in enumerate(values)
             if v is not None and not (isinstance(v, float) and math.isnan(v))]
    result = [(None, None)] * n
    if not valid:
        return result
    total = len(valid)
    sorted_valid = sorted(valid, key=lambda x: x[1], reverse=True)
    # Excel RANK: tied values get same rank (highest position = RANK.EQ)
    ranks = [None] * n
    rank = 1
    for pos, (idx, val) in enumerate(sorted_valid):
        if pos > 0 and val == sorted_valid[pos - 1][1]:
            pass  # tie: same rank
        else:
            rank = pos + 1
        ranks[idx] = rank

    # Excel QUARTILE = PERCENTILE.INC: position = q * (n-1) / 4
    sorted_ranks = sorted(r for r in ranks if r is not None)
    def quartile(q):
        pos = q * (total - 1) / 4
        lo = int(math.floor(pos))
        hi = int(math.ceil(pos))
        if lo == hi:
            return sorted_ranks[lo]
        frac = pos - lo
        return sorted_ranks[lo] + frac * (sorted_ranks[hi] - sorted_ranks[lo])

    t1 = quartile(1)
    t2 = quartile(2)
    t3 = quartile(3)

    # Single fund group: always Q1
    if total == 1:
        idx = valid[0][0]
        result[idx] = (1, 1)
        return result

    for i, r in enumerate(ranks):
        if r is None:
            continue
        if r < t1:
            q = 1
        elif r < t2:
            q = 2
        elif r < t3:
            q = 3
        else:
            q = 4
        result[i] = (r, q)
    return result


def _fetch_scoring_by_name():
    """Fetch scoring + daily NAV API, return {norm(name): scoring_record}."""
    from fetch_investdata import get_token
    import requests

    token = get_token()
    headers = {"Authorization": f"Bearer {token}"}

    # Fetch scoring data
    r = requests.get("https://api.infovesta.com/api/mutualfund/data/scoring", headers=headers)
    r.raise_for_status()
    scoring_data = r.json()

    # Fetch daily NAV for name → productId mapping
    r = requests.get("https://api.infovesta.com/api/mutualfund/data/dailynavlatesttwodates", headers=headers)
    r.raise_for_status()
    nav_data = r.json()

    # Build productId → name mapping
    name_by_pid = {}
    for item in nav_data:
        name_by_pid[item["productId"]] = item["name"]

    # Build normalized name → scoring record
    result = {}
    for s in scoring_data:
        pid = s["productId"]
        display_name = name_by_pid.get(pid)
        if not display_name:
            continue
        result[norm(display_name)] = s
        # Also index by stripped TR name
        tr_stripped = strip_tr(display_name)
        if tr_stripped != display_name:
            result[norm(tr_stripped)] = s

    return result


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def process(input_path: str, output_path: str,
            universe_path: str = "fund_universe.json",
            fund_map_path: str = "fund_map.json"):
    global FUND_MAP
    FUND_MAP = load_fund_map(fund_map_path)
    universe = load_universe(universe_path)

    print(f"Fund map overrides : {len(FUND_MAP)}")
    print(f"Reading            : {input_path}")

    # Extract NAV date from D-1 header row
    _wb_tmp = openpyxl.load_workbook(input_path, data_only=True)
    _d1_header = str(_wb_tmp["D-1"].cell(1, 1).value or "")
    nav_date_str = _d1_header.replace("Data Per Tanggal : ", "").strip() if "Tanggal" in _d1_header else None
    _wb_tmp.close()

    xl = pd.ExcelFile(input_path)
    d1       = pd.read_excel(xl, "D-1",       header=1)
    d2       = pd.read_excel(xl, "D-2",       header=1)
    brimi_d1 = pd.read_excel(xl, "BRIMI D-1", header=1)
    brimi_d2 = pd.read_excel(xl, "BRIMI D-2", header=1)
    indeks   = pd.read_excel(xl, "INDEKS",    header=1)
    indeks   = indeks.rename(columns={indeks.columns[0]: "Nama", indeks.columns[1]: "Nilai"})

    # Dynamic AUM column detection (name contains "AUM" and a date)
    aum_col = next((c for c in d1.columns if "AUM" in str(c) and "MI" not in str(c)
                    and "Shared" not in str(c)), None)
    if aum_col:
        D1_COL["AUM"] = aum_col
        print(f"AUM column         : {aum_col}")
    else:
        print("WARNING: AUM column not found in D-1")

    d1_lkp       = build_lookup(d1)
    d2_lkp       = build_lookup(d2)
    brimi_d1_lkp = build_lookup(brimi_d1)
    brimi_d2_lkp = build_lookup(brimi_d2)
    indeks_lkp   = build_lookup(indeks)

    wb_in          = openpyxl.load_workbook(input_path, data_only=True)
    bloomberg_data = load_bloomberg(wb_in)

    pt_cfg  = get_page_table_config(universe)
    ppt_cfg = get_perf_table_config(universe)
    print(f"Page table         : {len(pt_cfg)} entries")
    print(f"Performance table  : {len(ppt_cfg)} entries")

    # ── Page Table ────────────────────────────────────────────────────────
    pt_rows = []
    for f in pt_cfg:
        perf = get_perf(f["display_name"], f.get("alias"), f.get("source", "d1"),
                        d1_lkp, d2_lkp, indeks_lkp, bloomberg_data)
        aum  = None if f["is_index"] else get_aum(
            f["display_name"], f.get("alias"), f["is_usd"],
            brimi_d1_lkp, brimi_d2_lkp, d1_lkp, d2_lkp)
        pt_rows.append({**f, **perf, "AUM": aum})

    pt_df = pd.DataFrame(pt_rows)
    pt_df["Rank 1Y"] = None

    # Ranking: sections with rank_group share a combined ranking scope;
    # sections without rank_group are ranked independently by section_idx.
    non_idx = pt_df[~pt_df["is_index"]].copy()

    # 1) Cross-section groups: combine all funds with the same rank_group
    grouped_names = set()
    if "rank_group" in non_idx.columns:
        for rg, grp in non_idx[non_idx["rank_group"].notna()].groupby("rank_group", sort=False):
            # De-duplicate by display_name (take first occurrence's value)
            seen = {}
            for idx, row in grp.iterrows():
                nm = row["display_name"]
                if nm not in seen:
                    seen[nm] = (idx, row["1 Thn(%)"])
            # Rank the de-duplicated set
            items = list(seen.values())
            values = [v for _, v in items]
            rq = rank_quartile(values)
            rank_map = {}
            for (nm, (idx, val)), (rank, _) in zip(seen.items(), rq):
                rank_map[nm] = rank
            # Apply ranks back to ALL rows in the group (including duplicates)
            for idx, row in grp.iterrows():
                pt_df.loc[idx, "Rank 1Y"] = rank_map.get(row["display_name"])
            grouped_names.update(grp.index)

    # 2) Independent sections: rank each section_idx separately
    remaining = non_idx[~non_idx.index.isin(grouped_names)]
    for sidx, grp in remaining.groupby("section_idx", sort=False):
        rq = rank_quartile(grp["1 Thn(%)"].tolist())
        pt_df.loc[grp.index, "Rank 1Y"] = [r for r, _ in rq]

    # ── Performance Table ─────────────────────────────────────────────────
    ppt_rows = []
    for f in ppt_cfg:
        perf = get_perf(f["display_name"], None, "d1",
                        d1_lkp, d2_lkp, indeks_lkp, bloomberg_data,
                        is_price_return=f.get("is_price_return", False))
        aum  = None if f.get("is_benchmark") else get_aum(
            f["display_name"], f.get("brimi_alias"), f.get("is_usd", False),
            brimi_d1_lkp, brimi_d2_lkp, d1_lkp, d2_lkp)
        ppt_rows.append({**f, **perf, "AUM": aum})

    ppt_df = pd.DataFrame(ppt_rows)
    ppt_df["Score 1Y"] = None
    ppt_df["Quartile"] = None
    ppt_df["Quartile_3Bln"] = None
    ppt_df["Quartile_6Bln"] = None
    ppt_df["Quartile_YTD"] = None
    if "is_benchmark" not in ppt_df.columns:
        ppt_df["is_benchmark"] = False
    ppt_df["is_benchmark"] = ppt_df["is_benchmark"].fillna(False).astype(bool)

    # ── Quartile ranking using reference Excel groups ─────────────────────
    # The reference "NEW VS MI2" sheet defines 26 fund groups. Quartile
    # thresholds depend on ranking ALL funds in each group, not just our 28.
    # Build a lookup: norm(name) -> {rank, quartile} per period.
    REF_GROUPS_FILE = os.path.join(os.path.dirname(__file__), "quartile_group_mapping.json")
    if os.path.exists(REF_GROUPS_FILE):
        with open(REF_GROUPS_FILE) as f:
            _gm = json.load(f)
        _name_to_gi = _gm["name_to_group"]
        _group_funds = {int(k): v for k, v in _gm["group_funds"].items()}

        # Build D-1 lookup: norm(name) -> {period_col: value}
        _d1_by_name = {}
        for _, r in d1.iterrows():
            _d1_by_name[norm(str(r.iloc[0]))] = {
                "3 Bln(%)": r.get("3 Bln(%)"),
                "6 Bln(%)": r.get("6 Bln(%)"),
                "1 Thn(%)": r.get("1 Thn(%)"),
                "YTD(%)": r.get("YTD(%)"),
            }

        # Build D-2 lookup as fallback for funds missing from D-1
        # (e.g. Global/USD funds reported on previous day)
        _d2_by_name = {}
        for _, r in d2.iterrows():
            _d2_by_name[norm(str(r.iloc[0]))] = {
                "3 Bln(%)": r.get("3 Bln(%)"),
                "6 Bln(%)": r.get("6 Bln(%)"),
                "1 Thn(%)": r.get("1 Thn(%)"),
                "YTD(%)": r.get("YTD(%)"),
            }

        def _get_perf(name: str, col: str):
            """Get performance value from D-1, falling back to D-2."""
            val = _d1_by_name.get(name, {}).get(col)
            if val is not None and not (isinstance(val, float) and math.isnan(val)):
                return val
            val = _d2_by_name.get(name, {}).get(col)
            if val is not None and not (isinstance(val, float) and math.isnan(val)):
                return val
            return None

        _quartile_lookup = {}
        for gi, g_norm_names in _group_funds.items():
            for period_col in ["3 Bln(%)", "6 Bln(%)", "1 Thn(%)", "YTD(%)"]:
                items = [_get_perf(n, period_col) for n in g_norm_names]
                rq = rank_quartile(items)
                for n, (rank, q) in zip(g_norm_names, rq):
                    if rank is not None:
                        _quartile_lookup.setdefault(n, {})[period_col] = (rank, q)

        # Also index by TR → non-TR alias for funds that exist in D-1
        # without " - total return*" suffix. Only add if non-TR name doesn't
        # already have its own entry (both may be in the same group).
        for tr_name, entry in list(_quartile_lookup.items()):
            non_tr = tr_name.replace(" - total return*", "").strip()
            if non_tr and non_tr != tr_name and (non_tr in _d1_by_name or non_tr in _d2_by_name) and non_tr not in _quartile_lookup:
                _quartile_lookup[non_tr] = entry
    else:
        _quartile_lookup = {}

    # Apply quartile to our 28 BRI funds
    for idx, f in ppt_df.iterrows():
        if f.get("is_benchmark"):
            continue
        display = f["display_name"]
        # Try display_name first, then brimi_alias, then TR variant
        candidates = [norm(display)]
        alias = f.get("brimi_alias")
        if alias:
            candidates.append(norm(alias))
        # Also try appending " - total return*" for funds where reference uses TR name
        candidates.append(norm(display) + " - total return*")

        entry = {}
        for c in candidates:
            e = _quartile_lookup.get(c, {})
            if e:
                entry = e
                break

        for period_col, q_col in [
            ("3 Bln(%)", "Quartile_3Bln"),
            ("6 Bln(%)", "Quartile_6Bln"),
            ("1 Thn(%)", "Quartile"),
            ("YTD(%)", "Quartile_YTD"),
        ]:
            if period_col in entry:
                ppt_df.loc[idx, q_col] = entry[period_col][1]

    # Scoring API: Score 1Y from rankoneyearUrl
    scoring_lookup = _fetch_scoring_by_name()
    for idx, f in ppt_df.iterrows():
        if f.get("is_benchmark"):
            continue
        display = f["display_name"]
        scoring = scoring_lookup.get(norm(display))
        if scoring is None:
            scoring = scoring_lookup.get(norm(strip_tr(display)))

        if scoring:
            url = scoring.get("rankoneyearUrl", "")
            m = re.search(r"images/(.+)\.png$", url)
            if m:
                raw = m.group(1)
                ppt_df.loc[idx, "Score 1Y"] = SCORE_STARS.get(raw, raw)
            else:
                ppt_df.loc[idx, "Score 1Y"] = "No Score"
        else:
            ppt_df.loc[idx, "Score 1Y"] = "No Score"

    # Compute since_inception: (NAV / 1000 - 1) * 100
    ppt_df["since_inception"] = None
    for idx, f in ppt_df.iterrows():
        nav = f.get("NAB/UP")
        if nav is not None and not (isinstance(nav, float) and math.isnan(nav)):
            try:
                ppt_df.loc[idx, "since_inception"] = (float(nav) / 1000 - 1) * 100
            except (TypeError, ValueError):
                pass

    # ── Write ─────────────────────────────────────────────────────────────
    print(f"Writing            : {output_path}")
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)
    write_page_table(wb_out, pt_df, nav_date_str)
    write_perf_table(wb_out, ppt_df, nav_date_str)
    wb_out.save(output_path)
    print("Done.")


# ─────────────────────────────────────────────────────────────────────────────
# OUTPUT: NEW PAGE TABLE
# ─────────────────────────────────────────────────────────────────────────────

# ── Formatting constants (matched from original Excel) ────────────────────
FONT_NAME = "Nirmala Text"
FONT_DEFAULT = Font(name=FONT_NAME, size=11, color="000000")
FILL_NAV_DATE = PatternFill("solid", fgColor="E7E6E6")
FILL_COL_HDR  = PatternFill("solid", fgColor="1F4E78")
FILL_SECTION  = PatternFill("solid", fgColor="BDD7EE")
FILL_INDEX    = PatternFill("solid", fgColor="B4C6E7")
FILL_SPACER   = PatternFill("solid", fgColor="D0CECE")

FONT_NAV_DATE = Font(name=FONT_NAME, bold=True, size=10, color="000000")
FONT_TOP_GRP  = Font(name=FONT_NAME, bold=True, size=16, color="000000")
FONT_COL_HDR  = Font(name=FONT_NAME, bold=True, size=11, color="F2F2F2")
FONT_SECTION  = Font(name=FONT_NAME, size=11)  # not bold, no explicit color
FONT_BRI      = Font(name=FONT_NAME, bold=True, size=11, color="000000")
FONT_FUND     = Font(name=FONT_NAME, size=11, color="000000")
FONT_INDEX    = Font(name=FONT_NAME, bold=True, size=11, color="000000")


def top_group(section: str) -> str:
    s = section.lower()
    if any(k in s for k in ["equity", "global sharia equity", "index & etf", "etf fund"]):
        return "Equity Fund"
    if "money market" in s or "pasar uang" in s:
        return "Money Market Fund"
    if any(k in s for k in ["balanced", "berimbang", "campuran"]):
        return "Balanced Fund"
    if any(k in s for k in ["fixed income", "pendapatan tetap", "obligasi"]):
        return "Fixed Income Fund"
    return section


def _fill_row(ws, row: int, fill, font, ncols: int = 13):
    """Apply fill + font to columns A..M of a row."""
    for c in range(1, 1 + ncols):
        ws.cell(row, c).fill = fill
        ws.cell(row, c).font = font


def _is_bri(name: str) -> bool:
    """Check if a fund name is a BRI fund."""
    return str(name).upper().startswith("BRI ")


PT_DATA_COLS = ["1 Hr(%)", "1 Mgg(%)", "MTD(%)", "1 Bln(%)", "3 Bln(%)",
                "6 Bln(%)", "YTD(%)", "1 Thn(%)", "3 Thn(%)"]
PT_HEADERS   = ["Nama", "NAB/UP", "1 Hr (%)", "1 Mgg (%)", "MTD (%)",
                "1 Bln (%)", "3 Bln (%)", "6 Bln (%)", "YTD (%)",
                "1 Thn (%)", "3 Thn (%)", "Rank 1Y"]


def write_page_table(wb, df, nav_date_str: str | None = None):
    ws = wb.create_sheet("NEW PAGE TABLE (output)")
    NCOLS = 13  # A..M
    row = 1

    # ── Row 1: empty
    row += 1

    # ── Row 2: *NAV as of …  (use date from source data)
    nav_label = f"*NAV as of {nav_date_str}" if nav_date_str else "*NAV as of (date unavailable)"
    ws.cell(row, 1, nav_label)
    _fill_row(ws, row, FILL_NAV_DATE, FONT_NAV_DATE, ncols=NCOLS)
    row += 1

    # ── Rows 3-4: empty spacer
    row += 2

    prev_top = None
    for sidx, grp in df.groupby("section_idx", sort=False):
        section = grp.iloc[0]["section"]
        top = top_group(section)

        # ── Top group header (e.g., "Equity Fund") — only when group changes
        if top != prev_top:
            ws.cell(row, 1, top).font = FONT_TOP_GRP
            # Apply bold to all cols but only A gets size 16
            for c in range(2, 1 + NCOLS):
                ws.cell(row, c).font = Font(name=FONT_NAME, bold=True, size=11)
            row += 1
            prev_top = top

        # ── Column header row
        is_usd = bool(grp.iloc[0]["is_usd"])
        aum_lbl = f"AUM dalam {'Juta USD' if is_usd else 'Rp Miliar'}"
        for ci, h in enumerate(PT_HEADERS + [aum_lbl], 1):
            c = ws.cell(row, ci, h)
            c.alignment = Alignment(horizontal="center")
        _fill_row(ws, row, FILL_COL_HDR, FONT_COL_HDR, ncols=NCOLS)
        row += 1

        # ── Section name row (e.g., "Equity Fund (Big Cap)")
        ws.cell(row, 1, section)
        _fill_row(ws, row, FILL_SECTION, FONT_SECTION, ncols=NCOLS)
        row += 1

        # ── Fund data rows — sorted: non-index by rank ASC, then indexes
        non_idx = grp[~grp["is_index"]].copy()
        idx_rows = grp[grp["is_index"]].copy()

        # Sort non-index funds by rank (ascending). Funds without rank go last.
        if len(non_idx) > 0 and "Rank 1Y" in non_idx.columns:
            non_idx = non_idx.copy()
            non_idx["_sort_rank"] = non_idx["Rank 1Y"].apply(
                lambda x: x if x is not None and not (isinstance(x, float) and math.isnan(x)) else 9999
            )
            non_idx = non_idx.sort_values("_sort_rank", kind="mergesort")

        for _, f in non_idx.iterrows():
            row = _write_fund_row(ws, row, f, NCOLS)

        # ── Spacer row before indexes (gray D0CECE, empty)
        if len(idx_rows) > 0:
            _fill_row(ws, row, FILL_SPACER, FONT_FUND, ncols=NCOLS)
            row += 1

        # ── Index rows (blue B4C6E7 background)
        for _, f in idx_rows.iterrows():
            row = _write_index_row(ws, row, f, NCOLS)

        # ── Empty row after section
        row += 1

    # ── Column widths (match original)
    ws.column_dimensions["A"].width = 62.16
    ws.column_dimensions["B"].width = 13.66
    ws.column_dimensions["C"].width = 13.66
    for letter in "DEFGHIJK":
        ws.column_dimensions[letter].width = 13.0
    ws.column_dimensions["L"].width = 19.0
    ws.column_dimensions["M"].width = 20.33
    ws.freeze_panes = "A5"


def _write_fund_row(ws, row: int, f, ncols: int) -> int:
    """Write a single fund data row. Returns next row number."""
    rank = f.get("Rank 1Y")
    vals = (
        [f["display_name"], _fmt(f.get("NAB/UP"))]
        + [_fmt(f.get(c)) for c in PT_DATA_COLS]
        + [int(rank) if rank is not None and not (isinstance(rank, float) and math.isnan(rank)) else None,
           _fmt(f.get("AUM"))]
    )
    for ci, v in enumerate(vals, 1):
        cell = ws.cell(row, ci, v)
        if ci > 1:
            cell.alignment = Alignment(horizontal="right")
        if isinstance(v, float) and ci > 2:
            cell.number_format = "#,##0.0000"
    # BRI funds get bold
    if _is_bri(f["display_name"]):
        for c in range(1, 1 + ncols):
            ws.cell(row, c).font = FONT_BRI
    else:
        for c in range(1, 1 + ncols):
            ws.cell(row, c).font = FONT_FUND
    return row + 1


def _write_index_row(ws, row: int, f, ncols: int) -> int:
    """Write an index/benchmark row with blue B4C6E7 background."""
    vals = (
        [f["display_name"], _fmt(f.get("NAB/UP"))]
        + [_fmt(f.get(c)) for c in PT_DATA_COLS]
        + [None, None]  # No rank, no AUM for indexes
    )
    for ci, v in enumerate(vals, 1):
        cell = ws.cell(row, ci, v)
        if ci > 1:
            cell.alignment = Alignment(horizontal="right")
        if isinstance(v, float) and ci > 2:
            cell.number_format = "#,##0.0000"
    _fill_row(ws, row, FILL_INDEX, FONT_INDEX, ncols=ncols)
    return row + 1


# ─────────────────────────────────────────────────────────────────────────────
# OUTPUT: NEW PERFORMANCE TABLE
# ─────────────────────────────────────────────────────────────────────────────

PPT_DATA_COLS = ["1 Hr(%)", "1 Mgg(%)", "1 Bln(%)", "3 Bln(%)",
                 "6 Bln(%)", "YTD(%)", "1 Thn(%)", "3 Thn(%)", "5 Thn(%)"]
PPT_HEADERS   = ["Nama", "NAB/UP", "1 Hr (%)", "1 Mgg (%)", "1 Bln (%)",
                 "3 Bln (%)", "6 Bln (%)", "YTD (%)", "1 Thn (%)", "3 Thn (%)",
                 "5 Thn (%)", "Since Inception", "SI Date", "AUM (Rp Miliar)",
                 "Score 1Y", "Quartile 3Bln", "Quartile 6Bln", "Quartile 1Y",
                 "Quartile YTD"]


def write_perf_table(wb, df, nav_date_str=None):
    """Write NEW PERFORMANCE TABLE (output) sheet with all columns."""
    ws  = wb.create_sheet("NEW PERFORMANCE TABLE (output)")
    row = 1
    CS  = 4   # start at column D
    NC  = len(PPT_HEADERS)

    ws.cell(row, CS, f"BRI MI Daily Performance  |  Data Per {datetime.today().strftime('%d %B %Y')}")
    ws.cell(row, CS).font = Font(bold=True, size=12)
    row += 2

    # Row 6: main headers
    for ci, h in enumerate(PPT_HEADERS, CS):
        c = ws.cell(row, ci, h)
        c.font      = Font(bold=True, color="FFFFFF", size=9)
        c.fill      = PatternFill("solid", fgColor="1F4E79")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[row].height = 32
    row += 1

    # Row 7: sub-headers under Quartile columns
    for ci, label in enumerate(["3 Bln", "6 Bln", "1Y", "YTD"], CS + NC - 4):
        c = ws.cell(row, ci, label)
        c.font      = Font(bold=True, color="FFFFFF", size=8)
        c.fill      = PatternFill("solid", fgColor="1F4E79")
        c.alignment = Alignment(horizontal="center")
    row += 1

    prev_group = None
    for _, f in df.iterrows():
        group = f.get("group", "")

        if group != prev_group:
            if prev_group is not None:
                row += 1
            ws.cell(row, CS, group)
            for c in range(CS, CS + NC):
                ws.cell(row, c).fill = PatternFill("solid", fgColor="2E75B6")
                ws.cell(row, c).font = Font(bold=True, color="FFFFFF")
            row += 1
            prev_group = group

        # Quartile labels
        q_3bln  = f.get("Quartile_3Bln")
        q_6bln  = f.get("Quartile_6Bln")
        q_1y    = f.get("Quartile")
        q_ytd   = f.get("Quartile_YTD")
        q_lbl = lambda v: QUARTILE_LABELS.get(int(v)) if v is not None and not (isinstance(v, float) and math.isnan(v)) else None

        # Score 1Y: star text format from scoring API (e.g. "* * * -", "* +", "No Score")
        score_str = f.get("Score 1Y")

        # SI Date
        si_date = f.get("si_date", "")

        vals = (
            [f["display_name"], _fmt(f.get("NAB/UP"))]
            + [_fmt(f.get(c)) for c in PPT_DATA_COLS]
            + [_fmt(f.get("since_inception")),
               si_date if si_date else None,
               _fmt(f.get("AUM")),
               score_str,
               q_lbl(q_3bln), q_lbl(q_6bln), q_lbl(q_1y), q_lbl(q_ytd)]
        )
        for ci, v in enumerate(vals, CS):
            cell = ws.cell(row, ci, v)
            if ci > CS and isinstance(v, float):
                cell.number_format = "#,##0.0000"
                cell.alignment     = Alignment(horizontal="right")

        if f.get("is_benchmark"):
            for c in range(CS, CS + NC):
                ws.cell(row, c).fill = PatternFill("solid", fgColor="F2F2F2")
                ws.cell(row, c).font = Font(italic=True, color="666666")
        elif str(f["display_name"]).startswith("BRI "):
            for c in range(CS, CS + NC):
                ws.cell(row, c).fill = PatternFill("solid", fgColor="EBF3FB")

        # Color quartile cells
        for qi, q_val in enumerate([q_3bln, q_6bln, q_1y, q_ytd]):
            q_label = q_lbl(q_val)
            if q_label in QUARTILE_COLORS:
                qc = ws.cell(row, CS + NC - 4 + qi)
                qc.fill      = PatternFill("solid", fgColor=QUARTILE_COLORS[q_label])
                qc.font      = Font(bold=True, color="FFFFFF", size=9)
                qc.alignment = Alignment(horizontal="center")

        row += 1

    ws.column_dimensions["D"].width = 52
    ws.column_dimensions["E"].width = 12
    for ci in range(6, CS + NC):
        ws.column_dimensions[get_column_letter(ci)].width = 12
    ws.freeze_panes = "D6"


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    inp      = sys.argv[1]
    out      = sys.argv[2] if len(sys.argv) > 2 else "brimi_output.xlsx"
    universe = sys.argv[3] if len(sys.argv) > 3 else "fund_universe.json"
    fmap     = sys.argv[4] if len(sys.argv) > 4 else "fund_map.json"
    process(inp, out, universe, fmap)
