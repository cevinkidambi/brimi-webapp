"""
brimi_engine.py — Pipeline orchestrator for the BRIMI web app.

Coordinates the full processing flow:
  1. Fetch D-1/D-2/INDEKS data from Investdata API
  2. Build compiled Excel workbook from all 6 sources
  3. Run process_brimi.process() to produce output
  4. Return path to output file
"""

import os
import sys
import tempfile
import json
import re
import openpyxl
import pandas as pd

# Ensure the app's own directory is on the path so we can import sibling modules
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from fetch_investdata import (
    get_token, fetch_all, build_download_sheet, build_indeks_sheet,
    write_download_excel, write_indeks_excel, date_label, load_env
)
from raw_to_compiled import build_brimi_sheet, load_fund_map, norm, strip_tr


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TMP_DIR = os.path.join(BASE_DIR, "tmp")


def _ensure_tmp():
    os.makedirs(TMP_DIR, exist_ok=True)


def fetch_api_data(token, output_dir):
    """Fetch D-1, D-2, and INDEKS data from the API, write to output_dir."""
    print("Fetching data from Investdata API...")
    nav_data, aum_data, scoring_data, indeks_data = fetch_all(token)
    print(f"  NAV records: {len(nav_data)}")
    print(f"  AUM records: {len(aum_data)}")
    print(f"  Scoring records: {len(scoring_data)}")
    print(f"  Indeks records: {len(indeks_data)}")

    dates = sorted(set(r["date"] for r in nav_data), reverse=True)
    if len(dates) < 2:
        raise RuntimeError(f"Need at least 2 dates from API, got {dates}")
    d1_date = dates[0]
    d2_date = dates[1]
    print(f"  D-1 date: {d1_date}")
    print(f"  D-2 date: {d2_date}")

    d1_df = build_download_sheet(nav_data, aum_data, scoring_data, d1_date)
    d2_df = build_download_sheet(nav_data, aum_data, scoring_data, d2_date)
    indeks_df = build_indeks_sheet(indeks_data)

    d1_label = date_label(d1_date)
    d2_label = date_label(d2_date)
    indeks_label = d1_label

    d1_path = os.path.join(output_dir, f"Download Data Reksadana Per {d1_label}.xlsx")
    d2_path = os.path.join(output_dir, f"Download Data Reksadana Per {d2_label}.xlsx")
    indeks_path = os.path.join(output_dir, f"Indeks_{indeks_label}.xlsx")

    write_download_excel(d1_df, d1_path, f"Data Per Tanggal : {d1_date}")
    write_download_excel(d2_df, d2_path, f"Data Per Tanggal : {d2_date}")
    write_indeks_excel(indeks_df, indeks_path, f"Data Per {d1_date}")

    print(f"  Wrote D-1: {d1_path} ({len(d1_df)} funds)")
    print(f"  Wrote D-2: {d2_path} ({len(d2_df)} funds)")
    print(f"  Wrote INDEKS: {indeks_path} ({len(indeks_df)} entries)")

    return d1_path, d2_path, indeks_path


def build_compiled(d1_file, d2_file, nav_d1_file, nav_d2_file,
                   indeks_file, bb_file, output_path):
    """Build compiled Excel workbook from 6 raw source files.

    Adapted from raw_to_compiled.py's build_compiled(), using no reference Excel.
    """
    print(f"D-1 source      : {os.path.basename(d1_file)}")
    print(f"D-2 source      : {os.path.basename(d2_file)}")
    print(f"BRIMI D-1 source: {os.path.basename(nav_d1_file)}")
    print(f"BRIMI D-2 source: {os.path.basename(nav_d2_file)}")
    print(f"INDEKS source   : {os.path.basename(indeks_file)}")
    print(f"Bloomberg source: {os.path.basename(bb_file)}")

    d1_df = pd.read_excel(d1_file, header=1)
    d2_df = pd.read_excel(d2_file, header=1)
    nav_d1_df = pd.read_excel(nav_d1_file, header=1)
    nav_d2_df = pd.read_excel(nav_d2_file, header=1)

    fund_map = load_fund_map(os.path.join(BASE_DIR, "fund_map.json"))
    print(f"Fund map entries: {len(fund_map)}")

    # No reference — use HistoricalNAV names directly
    ref_d1 = nav_d1_df[["Fund Name"]].rename(columns={"Fund Name": "Nama"})
    ref_d2 = nav_d2_df[["Fund Name"]].rename(columns={"Fund Name": "Nama"})
    print(f"No reference: using HistoricalNAV names ({len(ref_d1)} funds)")

    brimi_d1_df = build_brimi_sheet(ref_d1, nav_d1_df, fund_map, fallback_df=d1_df)
    brimi_d2_df = build_brimi_sheet(ref_d2, nav_d2_df, fund_map, fallback_df=d2_df)
    print(f"BRIMI D-1: {len(brimi_d1_df)} funds")
    print(f"BRIMI D-2: {len(brimi_d2_df)} funds")

    # Extract date from D-1 header
    wb_hdr = openpyxl.load_workbook(d1_file, data_only=True)
    ws_hdr = wb_hdr["sheet 0"]
    nav_date = str(ws_hdr.cell(1, 1).value or "").replace("Data Per Tanggal : ", "").strip()
    wb_hdr.close()

    indeks_df = pd.read_excel(indeks_file, header=1)
    indeks_df = indeks_df.rename(
        columns={indeks_df.columns[0]: "Nama", indeks_df.columns[1]: "Nilai"}
    )

    wb_bb_src = openpyxl.load_workbook(bb_file, data_only=True)
    ws_bb_src = wb_bb_src["Sheet1"]

    print(f"Writing: {output_path}")
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    # D-1
    ws_d1 = wb_out.create_sheet("D-1")
    ws_d1.cell(1, 1, f"Data Per Tanggal : {nav_date}")
    for ci, col in enumerate(d1_df.columns, 1):
        ws_d1.cell(2, ci, col)
    for ri, row in enumerate(d1_df.values, 3):
        for ci, val in enumerate(row, 1):
            ws_d1.cell(ri, ci, val)
    print(f"D-1: {len(d1_df)} funds")

    # D-2
    ws_d2 = wb_out.create_sheet("D-2")
    ws_d2.cell(1, 1, "Data Per Tanggal : (D-2)")
    for ci, col in enumerate(d2_df.columns, 1):
        ws_d2.cell(2, ci, col)
    for ri, row in enumerate(d2_df.values, 3):
        for ci, val in enumerate(row, 1):
            ws_d2.cell(ri, ci, val)
    print(f"D-2: {len(d2_df)} funds")

    # BRIMI D-1
    ws_bd1 = wb_out.create_sheet("BRIMI D-1")
    for ci in range(1, 7):
        ws_bd1.cell(1, ci, "")
    for ci, col in enumerate(brimi_d1_df.columns, 1):
        ws_bd1.cell(2, ci, col)
    for ri, row in enumerate(brimi_d1_df.values, 3):
        for ci, val in enumerate(row, 1):
            ws_bd1.cell(ri, ci, val)

    # BRIMI D-2
    ws_bd2 = wb_out.create_sheet("BRIMI D-2")
    for ci in range(1, 7):
        ws_bd2.cell(1, ci, "")
    for ci, col in enumerate(brimi_d2_df.columns, 1):
        ws_bd2.cell(2, ci, col)
    for ri, row in enumerate(brimi_d2_df.values, 3):
        for ci, val in enumerate(row, 1):
            ws_bd2.cell(ri, ci, val)

    # INDEKS
    ws_in = wb_out.create_sheet("INDEKS")
    ws_in.cell(1, 1, f"Data Per {nav_date}")
    for ci, col in enumerate(indeks_df.columns, 1):
        ws_in.cell(2, ci, col)
    for ri, row in enumerate(indeks_df.values, 3):
        for ci, val in enumerate(row, 1):
            ws_in.cell(ri, ci, val)
    print(f"INDEKS: {len(indeks_df)} entries")

    # BloombergIndex
    ws_bb = wb_out.create_sheet("BloombergIndex")
    for row_idx in range(1, ws_bb_src.max_row + 1):
        for col_idx in range(1, ws_bb_src.max_column + 1):
            src_cell = ws_bb_src.cell(row_idx, col_idx)
            ws_bb.cell(row_idx, col_idx).value = src_cell.value

    wb_out.save(output_path)
    wb_bb_src.close()
    print("Compiled workbook built successfully.")


def run_pipeline(historicalnav_t1_path, historicalnav_t2_path, bloomberg_path,
                 log_callback=None):
    """Run the full BRIMI pipeline.

    Args:
        historicalnav_t1_path: path to HistoricalNAV T-1 file (newest)
        historicalnav_t2_path: path to HistoricalNAV T-2 file (previous)
        bloomberg_path: path to indeks_bloomberg.xlsx
        log_callback: optional callable(str) for streaming log output

    Returns:
        tuple: (output_path, nav_date_str)
    """
    _ensure_tmp()

    def _log(msg):
        print(msg)
        if log_callback:
            log_callback(msg)

    try:
        # Step 1: Fetch API data
        token = get_token()
        api_dir = TMP_DIR
        d1_path, d2_path, indeks_path = fetch_api_data(token, api_dir)

        # Step 2: Build compiled workbook
        compiled_path = os.path.join(TMP_DIR, "compiled_input.xlsx")
        build_compiled(d1_path, d2_path, historicalnav_t1_path,
                       historicalnav_t2_path, indeks_path, bloomberg_path,
                       compiled_path)

        # Step 3: Run process_brimi
        output_path = os.path.join(TMP_DIR, "brimi_output.xlsx")
        from process_brimi import process
        process(compiled_path, output_path,
                os.path.join(BASE_DIR, "fund_universe.json"),
                os.path.join(BASE_DIR, "fund_map.json"))

        # Extract date from compiled D-1 for download filename
        wb_tmp = openpyxl.load_workbook(compiled_path, data_only=True)
        header = str(wb_tmp["D-1"].cell(1, 1).value or "")
        nav_date = header.replace("Data Per Tanggal : ", "").strip()
        wb_tmp.close()

        _log(f"\nDone. Output: {output_path}")
        return output_path, nav_date

    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        _log(f"\nERROR: {e}")
        _log(tb)
        raise
