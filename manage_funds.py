"""
manage_funds.py — Fund Universe Manager
========================================
Add, remove, deactivate, and list funds in fund_universe.json without
touching any Python code or Excel files.

Commands
--------
  --init <excel>               Seed fund_universe.json from an existing Excel template
  --list                       Print all active funds (page table + perf table)
  --list-all                   Include inactive funds
  --add-page <section> <name>  Add a fund to a section in the page table
  --add-perf <group> <name>    Add a fund to the performance table
  --remove <name>              Deactivate a fund everywhere (sets active=false)
  --delete <name>              Permanently delete a fund everywhere
  --reactivate <name>          Re-activate a previously deactivated fund
  --rename <old> <new>         Rename a fund everywhere
  --alias <name> <alias>       Set/update the alias for a fund (page table)
  --brimi-alias <name> <alias> Set/update the brimi_alias for a fund (perf table)
  --sections                   List all section names in the page table

Options
-------
  --universe <path>   Path to fund_universe.json  (default: fund_universe.json)
  --excel <path>      Path to input Excel (only needed for --init)

Examples
--------
  python manage_funds.py --init "27_April_2026_BRIMI.xlsx"
  python manage_funds.py --list
  python manage_funds.py --sections
  python manage_funds.py --add-page "Equity Fund (Big Cap)" "Trimegah Dana Saham"
  python manage_funds.py --add-perf "Equity" "BRI Mawar Baru"
  python manage_funds.py --remove "BNP Paribas Pesona"
  python manage_funds.py --reactivate "BNP Paribas Pesona"
  python manage_funds.py --alias "BRI Mawar" "BRI Mawar - Total Return*"
  python manage_funds.py --brimi-alias "BRI Seruni Likuid Dolar" "BRI Seruni Likuid Dolar"
  python manage_funds.py --rename "Old Fund Name" "New Fund Name"
"""

import sys
import re
import json
import os
import argparse

UNIVERSE_DEFAULT = "fund_universe.json"
INDEKS_NAMES = {
    "LQ45", "IDX Small-Mid Cap Liquid", "Indeks Harga Saham Gabungan",
    "Indeks Saham Syariah Indonesia", "Infovesta sharia Equity fund Index",
    "Jakarta Islamic Index", "Infovesta Money Market Fund Index",
    "Infovesta Balanced Fund Index", "Infovesta Government Bond Index",
    "Infovesta Government Bond Index Short", "Infovesta Sharia Balanced Fund Index",
    "Infovesta Sharia Money Market Fund", "Infovesta Fixed Income Fund Index",
    "MSCI Indonesia ESG Screened", "Papan Akselerasi",
}
BLOOMBERG_NAMES = {"MSCI World Islamic Index", "MSCI Indonesia ESG Screened",
                   "INDOBex Government Total Return*"}
D2_SECTION_KEYWORDS = ["global sharia equity", "money market usd", "fixed income usd"]
SKIP_ALIAS = {"manual", "ganti sebulan sekali ( dari email)",
              "ganti sebulan sekali (fun fact sheet investdata)"}


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def norm(s: str) -> str:
    return str(s).strip().lower()


def load(path: str) -> dict:
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Fund universe not found: {path}\n"
            "Run: python manage_funds.py --init <excel>"
        )
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def save(universe: dict, path: str):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(universe, f, indent=2, ensure_ascii=False)
    print(f"Saved → {path}")


def find_in_page(universe: dict, name: str) -> list[tuple]:
    """Return [(section_idx, fund_idx), ...] for all matches."""
    hits = []
    for si, section in enumerate(universe["page_table"]["sections"]):
        for fi, fund in enumerate(section["funds"]):
            if norm(fund["display_name"]) == norm(name):
                hits.append((si, fi))
    return hits


def find_in_perf(universe: dict, name: str) -> list[int]:
    return [i for i, f in enumerate(universe["performance_table"]["funds"])
            if norm(f["display_name"]) == norm(name)]


# ─────────────────────────────────────────────────────────────────────────────
# COMMANDS
# ─────────────────────────────────────────────────────────────────────────────

def cmd_init(excel_path: str, universe_path: str):
    """Seed fund_universe.json from an existing Excel template."""
    try:
        import openpyxl
        import pandas as pd
    except ImportError:
        print("ERROR: openpyxl and pandas are required. pip install openpyxl pandas")
        sys.exit(1)

    if os.path.exists(universe_path):
        ans = input(f"{universe_path} already exists. Overwrite? [y/N] ").strip().lower()
        if ans != "y":
            print("Aborted.")
            return

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["NEW PAGE TABLE"]
    all_rows_pt = list(ws.iter_rows(min_row=1, max_row=400, values_only=True))

    # Names that are ONLY top-level group banners, never real sub-sections
    GROUP_ONLY = {"Equity Fund", "Money Market Fund", "Balanced Fund", "Fixed Income Fund"}
    # Sections with no raw data (manual/discretionary) — skip entirely
    REMOVE_SECS = {"Discretionary Fund", "LQ45"}

    def _next_nonempty_a(rows, from_idx):
        for j in range(from_idx + 1, len(rows)):
            a = rows[j][0]
            if a is not None and str(a).strip():
                return str(a).strip()
        return None

    sections = []
    current_is_usd = False
    current_aum_unit = "Rp Miliar"
    skip_current = False

    for i, row in enumerate(all_rows_pt):
        a = str(row[0]).strip() if row[0] is not None else ""
        b = row[1] if len(row) > 1 else None
        n = row[13] if len(row) > 13 else None
        m_hdr = row[12] if len(row) > 12 else None

        if b == "NAB/UP" and isinstance(m_hdr, str):
            current_aum_unit = "Juta USD" if "USD" in m_hdr else "Rp Miliar"

        if a and not isinstance(b, (int, float)) and b not in ("NAB/UP", "Apr-25") and a != "Nama":
            if b is None or (isinstance(b, str) and b.strip() in ("", " ")):
                if a.startswith("*") or a.startswith("PT "):
                    continue
                nxt_a = _next_nonempty_a(all_rows_pt, i)
                if a in GROUP_ONLY and nxt_a == "Nama":
                    continue  # top-level group banner, not a real section
                skip_current = (a in REMOVE_SECS)
                if skip_current:
                    continue
                current_is_usd = any(k in a.lower() for k in D2_SECTION_KEYWORDS)
                sections.append({"section": a, "is_usd": current_is_usd,
                                  "aum_unit": current_aum_unit, "funds": []})
                continue

        if skip_current:
            continue

        if a and isinstance(b, (int, float)) and a != "Nama" and sections:
            raw_alias = str(n).strip() if n else None
            alias = raw_alias if raw_alias and raw_alias.lower() not in SKIP_ALIAS else None
            is_index = a in INDEKS_NAMES or a in BLOOMBERG_NAMES
            sections[-1]["funds"].append({
                "display_name": a, "alias": alias,
                "is_index": is_index, "active": True,
            })

    # Performance table
    ws2 = wb["NEW PERFORMANCE PAGE "]
    GROUP_SEQ = ["Equity", "Balanced", "Fixed Income", "Money Market", "Other"]
    group_idx = 0
    separators_seen = 0
    perf_funds = []

    for row in ws2.iter_rows(min_row=1, max_row=50, values_only=True):
        col_c = row[2] if len(row) > 2 else None
        col_d = row[3] if len(row) > 3 else None
        col_e = row[4] if len(row) > 4 else None
        col_c_s = str(col_c).strip() if col_c else None
        col_d_s = str(col_d).strip() if col_d else None

        is_sep = (
            (col_d is None or not isinstance(col_e, (int, float)))
            and any(isinstance(row[j], str) and "VLOOKUP" in row[j]
                    for j in range(29, min(55, len(row))))
        )
        if is_sep:
            separators_seen += 1
            if separators_seen >= 2 and group_idx + 1 < len(GROUP_SEQ):
                group_idx += 1
            continue

        is_data = (col_d_s and col_d_s not in ("Nama", "nan", "None")
                   and not col_d_s.startswith("#") and not isinstance(col_d, (int, float))
                   and isinstance(col_e, (int, float)))
        if not is_data:
            continue

        raw = re.sub(r"\*+\s*$", "", col_d_s).strip()
        display = re.sub(r"\s*-\s*total return\s*$", "", raw, flags=re.IGNORECASE).strip()
        brimi_alias = col_c_s if col_c_s and col_c_s not in ("None", col_d_s) else None
        is_pr = bool(re.search(r"price return", raw, re.IGNORECASE))
        is_usd = any(k in display.lower() for k in ["dollar", "usd"])
        is_bench = display in INDEKS_NAMES or display in BLOOMBERG_NAMES

        perf_funds.append({
            "display_name": display, "brimi_alias": brimi_alias,
            "group": GROUP_SEQ[group_idx], "is_price_return": is_pr,
            "is_usd": is_usd, "is_benchmark": is_bench, "active": True,
        })

    universe = {
        "_comment": (
            "Fund universe config. Edit this file or use manage_funds.py to add/remove funds. "
            "Set active=false to temporarily exclude without deleting."
        ),
        "page_table": {
            "_comment": "Sections and funds shown in NEW PAGE TABLE output.",
            "sections": sections,
        },
        "performance_table": {
            "_comment": (
                "Funds in NEW PERFORMANCE TABLE. "
                "brimi_alias: AUM lookup key in BRIMI D-1/D-2 (null = use display_name). "
                "group: Equity | Balanced | Fixed Income | Money Market. "
                "is_price_return: true = use base NAV not Total Return."
            ),
            "funds": perf_funds,
        },
    }

    save(universe, universe_path)
    print(f"  Page table  : {len(sections)} sections, "
          f"{sum(len(s['funds']) for s in sections)} funds")
    print(f"  Perf table  : {len(perf_funds)} funds across "
          f"{len(set(f['group'] for f in perf_funds))} groups")


def cmd_list(universe: dict, include_inactive: bool = False):
    print("\n══ PAGE TABLE ══════════════════════════════════════════════")
    for section in universe["page_table"]["sections"]:
        funds = section["funds"] if include_inactive else [f for f in section["funds"] if f.get("active", True)]
        if not funds:
            continue
        print(f"\n  [{section['section']}]")
        for f in funds:
            status = "" if f.get("active", True) else "  ⚠ INACTIVE"
            alias  = f"  → alias: {f['alias']}" if f.get("alias") else ""
            idx    = "  [index]" if f.get("is_index") else ""
            print(f"    {f['display_name']}{alias}{idx}{status}")

    print("\n══ PERFORMANCE TABLE ════════════════════════════════════════")
    prev_group = None
    for f in universe["performance_table"]["funds"]:
        if not include_inactive and not f.get("active", True):
            continue
        if f["group"] != prev_group:
            print(f"\n  [{f['group']}]")
            prev_group = f["group"]
        status = "" if f.get("active", True) else "  ⚠ INACTIVE"
        ba = f"  → brimi_alias: {f['brimi_alias']}" if f.get("brimi_alias") else ""
        print(f"    {f['display_name']}{ba}{status}")


def cmd_sections(universe: dict):
    print("\nSections in PAGE TABLE:")
    for s in universe["page_table"]["sections"]:
        n = sum(1 for f in s["funds"] if f.get("active", True))
        print(f"  {s['section']!r:50s} — {n} active funds")


def cmd_add_page(universe: dict, section_name: str, fund_name: str,
                 alias: str | None, is_index: bool):
    sections = universe["page_table"]["sections"]
    matched = [s for s in sections if norm(s["section"]) == norm(section_name)]
    if not matched:
        print(f"ERROR: Section not found: {section_name!r}")
        print("Available sections:")
        for s in sections:
            print(f"  {s['section']}")
        return

    section = matched[0]
    if any(norm(f["display_name"]) == norm(fund_name) for f in section["funds"]):
        print(f"Fund already exists in section: {fund_name!r}. Use --reactivate if inactive.")
        return

    new_fund = {
        "display_name": fund_name,
        "alias": alias,
        "is_index": is_index,
        "active": True,
    }

    if is_index:
        # Benchmarks always go at the end
        section["funds"].append(new_fund)
    else:
        # Regular funds go before any existing benchmarks
        first_benchmark = next(
            (i for i, f in enumerate(section["funds"]) if f.get("is_index")),
            len(section["funds"])
        )
        section["funds"].insert(first_benchmark, new_fund)

    print(f"Added to page table [{section_name}]: {fund_name!r}"
          + (" [benchmark — placed at end]" if is_index else ""))


def cmd_add_perf(universe: dict, group: str, fund_name: str,
                 brimi_alias: str | None, is_price_return: bool, is_usd: bool):
    valid_groups = ["Equity", "Balanced", "Fixed Income", "Money Market", "Other"]
    if group not in valid_groups:
        print(f"ERROR: group must be one of: {valid_groups}")
        return

    funds = universe["performance_table"]["funds"]
    if any(norm(f["display_name"]) == norm(fund_name) for f in funds):
        print(f"Fund already exists in perf table: {fund_name!r}. Use --reactivate if inactive.")
        return

    # Insert at end of the correct group
    insert_at = len(funds)
    for i, f in enumerate(funds):
        if f["group"] == group:
            insert_at = i + 1

    is_bench = fund_name in INDEKS_NAMES or fund_name in BLOOMBERG_NAMES
    funds.insert(insert_at, {
        "display_name":    fund_name,
        "brimi_alias":     brimi_alias,
        "group":           group,
        "is_price_return": is_price_return,
        "is_usd":          is_usd,
        "is_benchmark":    is_bench,
        "active":          True,
    })
    print(f"Added to performance table [{group}]: {fund_name!r}")


def cmd_set_active(universe: dict, fund_name: str, active: bool):
    action = "Reactivated" if active else "Deactivated"
    found = False

    for section in universe["page_table"]["sections"]:
        for fund in section["funds"]:
            if norm(fund["display_name"]) == norm(fund_name):
                fund["active"] = active
                print(f"{action} in page table [{section['section']}]: {fund_name!r}")
                found = True

    for fund in universe["performance_table"]["funds"]:
        if norm(fund["display_name"]) == norm(fund_name):
            fund["active"] = active
            print(f"{action} in performance table [{fund['group']}]: {fund_name!r}")
            found = True

    if not found:
        print(f"Fund not found: {fund_name!r}")


def cmd_delete(universe: dict, fund_name: str):
    ans = input(f"Permanently delete {fund_name!r} everywhere? [y/N] ").strip().lower()
    if ans != "y":
        print("Aborted.")
        return

    deleted = False
    for section in universe["page_table"]["sections"]:
        before = len(section["funds"])
        section["funds"] = [f for f in section["funds"]
                            if norm(f["display_name"]) != norm(fund_name)]
        if len(section["funds"]) < before:
            print(f"Deleted from page table [{section['section']}]: {fund_name!r}")
            deleted = True

    before = len(universe["performance_table"]["funds"])
    universe["performance_table"]["funds"] = [
        f for f in universe["performance_table"]["funds"]
        if norm(f["display_name"]) != norm(fund_name)
    ]
    if len(universe["performance_table"]["funds"]) < before:
        print(f"Deleted from performance table: {fund_name!r}")
        deleted = True

    if not deleted:
        print(f"Fund not found: {fund_name!r}")


def cmd_rename(universe: dict, old_name: str, new_name: str):
    found = False
    for section in universe["page_table"]["sections"]:
        for fund in section["funds"]:
            if norm(fund["display_name"]) == norm(old_name):
                fund["display_name"] = new_name
                print(f"Renamed in page table [{section['section']}]: {old_name!r} → {new_name!r}")
                found = True
    for fund in universe["performance_table"]["funds"]:
        if norm(fund["display_name"]) == norm(old_name):
            fund["display_name"] = new_name
            print(f"Renamed in performance table [{fund['group']}]: {old_name!r} → {new_name!r}")
            found = True
    if not found:
        print(f"Fund not found: {old_name!r}")


def cmd_set_alias(universe: dict, fund_name: str, alias: str):
    found = False
    for section in universe["page_table"]["sections"]:
        for fund in section["funds"]:
            if norm(fund["display_name"]) == norm(fund_name):
                fund["alias"] = alias
                print(f"Set alias in page table [{section['section']}]: {fund_name!r} → {alias!r}")
                found = True
    if not found:
        print(f"Fund not found in page table: {fund_name!r}")


def cmd_set_brimi_alias(universe: dict, fund_name: str, alias: str):
    found = False
    for fund in universe["performance_table"]["funds"]:
        if norm(fund["display_name"]) == norm(fund_name):
            fund["brimi_alias"] = alias
            print(f"Set brimi_alias in perf table [{fund['group']}]: {fund_name!r} → {alias!r}")
            found = True
    if not found:
        print(f"Fund not found in performance table: {fund_name!r}")


# ─────────────────────────────────────────────────────────────────────────────
# ARGUMENT PARSING
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Manage fund_universe.json",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--universe", default=UNIVERSE_DEFAULT,
                        help="Path to fund_universe.json")
    parser.add_argument("--init", metavar="EXCEL",
                        help="Seed fund_universe.json from Excel template")
    parser.add_argument("--list", action="store_true",
                        help="List all active funds")
    parser.add_argument("--list-all", action="store_true",
                        help="List all funds including inactive")
    parser.add_argument("--sections", action="store_true",
                        help="List section names")
    parser.add_argument("--add-page", nargs=2, metavar=("SECTION", "NAME"),
                        help="Add a fund to page table section")
    parser.add_argument("--add-perf", nargs=2, metavar=("GROUP", "NAME"),
                        help="Add a fund to performance table group")
    parser.add_argument("--remove", metavar="NAME",
                        help="Deactivate a fund (set active=false)")
    parser.add_argument("--delete", metavar="NAME",
                        help="Permanently delete a fund")
    parser.add_argument("--reactivate", metavar="NAME",
                        help="Re-activate a deactivated fund")
    parser.add_argument("--rename", nargs=2, metavar=("OLD", "NEW"),
                        help="Rename a fund everywhere")
    parser.add_argument("--alias", nargs=2, metavar=("NAME", "ALIAS"),
                        help="Set alias for a fund in page table")
    parser.add_argument("--brimi-alias", nargs=2, metavar=("NAME", "ALIAS"),
                        help="Set brimi_alias for a fund in performance table")
    # Extra flags for --add-page / --add-perf
    parser.add_argument("--is-index", action="store_true",
                        help="Mark fund as index/benchmark (used with --add-page)")
    parser.add_argument("--page-alias", metavar="ALIAS",
                        help="Alias for the fund (used with --add-page)")
    parser.add_argument("--price-return", action="store_true",
                        help="Mark as price return lookup (used with --add-perf)")
    parser.add_argument("--usd", action="store_true",
                        help="Mark as USD fund (used with --add-perf)")
    parser.add_argument("--brimi-alias-val", metavar="ALIAS",
                        help="brimi_alias value (used with --add-perf)")

    args = parser.parse_args()

    if args.init:
        cmd_init(args.init, args.universe)
        return

    if args.list:
        cmd_list(load(args.universe))
        return

    if args.list_all:
        cmd_list(load(args.universe), include_inactive=True)
        return

    if args.sections:
        cmd_sections(load(args.universe))
        return

    # Commands that modify the file
    universe = load(args.universe)
    modified = False

    if args.add_page:
        cmd_add_page(universe, args.add_page[0], args.add_page[1],
                     alias=args.page_alias, is_index=args.is_index)
        modified = True

    if args.add_perf:
        cmd_add_perf(universe, args.add_perf[0], args.add_perf[1],
                     brimi_alias=args.brimi_alias_val,
                     is_price_return=args.price_return,
                     is_usd=args.usd)
        modified = True

    if args.remove:
        cmd_set_active(universe, args.remove, active=False)
        modified = True

    if args.delete:
        cmd_delete(universe, args.delete)
        modified = True

    if args.reactivate:
        cmd_set_active(universe, args.reactivate, active=True)
        modified = True

    if args.rename:
        cmd_rename(universe, args.rename[0], args.rename[1])
        modified = True

    if args.alias:
        cmd_set_alias(universe, args.alias[0], args.alias[1])
        modified = True

    if args.brimi_alias:
        cmd_set_brimi_alias(universe, args.brimi_alias[0], args.brimi_alias[1])
        modified = True

    if modified:
        save(universe, args.universe)
    elif not any([args.list, args.list_all, args.sections, args.init]):
        parser.print_help()


if __name__ == "__main__":
    main()
