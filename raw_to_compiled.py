"""
raw_to_compiled.py — Convert raw data files to compiled Excel format
=====================================================================
Reads files from very_raw_data/ and produces a compiled Excel file
with the exact sheet structure that process_brimi.py expects:
  D-1, D-2, BRIMI D-1, BRIMI D-2, INDEKS, BloombergIndex

BRIMI sheets: uses a reference compiled Excel for fund names (since
HistoricalNAV uses different naming conventions), then looks up AUM
values from HistoricalNAV using fund_map.json overrides + fuzzy matching.

When funds are added/removed weekly, update fund_universe.json and
fund_map.json as normal — this script uses those same configs.

Usage:
    python raw_to_compiled.py <output.xlsx> [reference_compiled.xlsx]

Raw data sources:
  Download Data Reksadana Per 27-Apr-2026.xlsx → D-1
  Download Data Reksadana Per 24-Apr-2026.xlsx → D-2
  HistoricalNAV_280426.xlsx                    → BRIMI D-1 (BRI AUM)
  HistoricalNAV_270426_2.xlsx                  → BRIMI D-2 (BRI AUM)
  Indeks_27-Apr-2026.xlsx                      → INDEKS
  indeks_bloomberg.xlsx                        → BloombergIndex
"""

import sys
import os
import re
import json
import pandas as pd
import openpyxl

RAW_DIR = "very_raw_data"


def norm(s: str) -> str:
    return str(s).strip().lower()


def strip_tr(name: str) -> str:
    return re.sub(r"\s*-\s*total return\*?\s*$", "", name, flags=re.IGNORECASE).strip()


def find_nav_row(brimi_name: str, fund_map: dict, nav_lower: dict) -> dict | None:
    """Find HistoricalNAV row matching a BRIMI fund name.

    Same priority as process_brimi.py get_aum():
    1. fund_map.json brimi_name override
    2. Exact match (case-insensitive)
    3. Match after stripping ' - Total Return*' suffix
    4. HistoricalNAV contains BRIMI name (substring)
    """
    override = fund_map.get(norm(brimi_name), {})
    brimi_override = override.get("brimi_name")

    candidates = []
    if brimi_override:
        candidates.append(norm(brimi_override))
    candidates.append(norm(brimi_name))
    br_no_tr = strip_tr(brimi_name)
    br_no_tr_norm = norm(br_no_tr)
    if br_no_tr_norm not in candidates:
        candidates.append(br_no_tr_norm)

    for candidate in candidates:
        if candidate in nav_lower:
            return nav_lower[candidate]

    # Substring: HistoricalNAV contains BRIMI name
    for candidate in candidates:
        if len(candidate) > 5:
            for nav_norm, nav_row in nav_lower.items():
                if candidate in nav_norm:
                    return nav_row

    return None


def load_fund_map(path="fund_map.json"):
    """Get {display_name_lower: {brimi_name, d1_name, nav_name}}."""
    if not os.path.exists(path):
        return {}
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    return {
        norm(e["display_name"]): {
            "brimi_name": e.get("brimi_name"),
            "d1_name": e.get("d1_name"),
            "nav_name": e.get("nav_name"),
        }
        for e in data.get("mappings", [])
    }


def build_brimi_sheet(ref_brimi_df: pd.DataFrame, nav_df: pd.DataFrame, fund_map: dict,
                      fallback_df: pd.DataFrame = None):
    """Build BRIMI-style sheet with names from reference, AUM from HistoricalNAV.

    Args:
        ref_brimi_df: BRIMI sheet from compiled Excel (has correct names)
        nav_df: HistoricalNAV data (has AUM values)
        fund_map: name override mappings
        fallback_df: Download Data for AUM fallback
    """
    nav_lower = {norm(r["Fund Name"]): r for _, r in nav_df.iterrows()}
    fallback_lookup = {}
    if fallback_df is not None:
        fallback_lookup = {norm(r["Nama"]): r for _, r in fallback_df.iterrows()}

    rows = []
    for _, ref_row in ref_brimi_df.iterrows():
        brimi_name = str(ref_row["Nama"]).strip()
        aum_value = 0

        nav_row = find_nav_row(brimi_name, fund_map, nav_lower)
        if nav_row is not None:
            aum_value = float(nav_row["AUM"]) if pd.notna(nav_row.get("AUM")) else 0

        # Fallback: if AUM is 0, try Download Data
        if aum_value == 0 and fallback_df is not None:
            fb_row = fallback_lookup.get(norm(brimi_name))
            if fb_row is not None:
                aum_col = next((c for c in fb_row.index if "AUM" in str(c) and "MI" not in str(c)), None)
                if aum_col and pd.notna(fb_row[aum_col]):
                    aum_value = float(fb_row[aum_col])

        rows.append({
            "Nama": brimi_name,
            "Jenis": "",
            "Manajer Investasi": "",
            "Kustodian": "",
            "Denominasi": "",
            "AUM": aum_value,
        })

    return pd.DataFrame(rows)


def find_files(raw_dir):
    """Find raw data files by glob pattern, sorted by date (newest first)."""
    import glob

    def find(pattern, label):
        matches = sorted(glob.glob(os.path.join(raw_dir, pattern)))
        if not matches:
            print(f"ERROR: No {label} files matching {pattern} in {raw_dir}")
            sys.exit(1)
        return matches[-1]  # newest (sorted by name)

    d1_file = find("Download Data Reksadana Per *.xlsx", "D-1")
    d2_files = [f for f in sorted(glob.glob(os.path.join(raw_dir, "Download Data Reksadana Per *.xlsx"))) if f != d1_file]
    if not d2_files:
        print(f"ERROR: Need at least 2 Download Data files, found 1")
        sys.exit(1)
    d2_file = d2_files[-1]  # second newest

    nav_files = sorted(glob.glob(os.path.join(raw_dir, "HistoricalNAV_*.xlsx")))
    if len(nav_files) < 2:
        print(f"ERROR: Need at least 2 HistoricalNAV files, found {len(nav_files)}")
        sys.exit(1)
    nav_28_file = nav_files[-1]  # newest
    nav_27_file = nav_files[-2]  # second newest

    indeks_files = sorted(glob.glob(os.path.join(raw_dir, "Indeks_*.xlsx")))
    if not indeks_files:
        indeks_files = sorted(glob.glob(os.path.join(raw_dir, "Indeks*.xlsx")))
    if not indeks_files:
        print(f"ERROR: No Indeks file found in {raw_dir}")
        sys.exit(1)
    indeks_file = indeks_files[-1]

    bb_file = os.path.join(raw_dir, "indeks_bloomberg.xlsx")
    if not os.path.exists(bb_file):
        print(f"ERROR: Missing Bloomberg file: {bb_file}")
        sys.exit(1)

    return d1_file, d2_file, nav_28_file, nav_27_file, indeks_file, bb_file


def build_compiled(output_path: str, reference_path: str = None):
    """Build compiled Excel from raw data files."""

    # Find raw files dynamically
    d1_file, d2_file, nav_28_file, nav_27_file, indeks_file, bb_file = find_files(RAW_DIR)

    for label, path in [("D-1", d1_file), ("D-2", d2_file),
                        ("BRIMI D-1", nav_28_file), ("BRIMI D-2", nav_27_file),
                        ("INDEKS", indeks_file), ("Bloomberg", bb_file)]:
        if not os.path.exists(path):
            print(f"ERROR: Missing {label} file: {path}")
            sys.exit(1)

    print(f"D-1 source      : {os.path.basename(d1_file)}")
    print(f"D-2 source      : {os.path.basename(d2_file)}")
    print(f"BRIMI D-1 source: {os.path.basename(nav_28_file)}")
    print(f"BRIMI D-2 source: {os.path.basename(nav_27_file)}")
    print(f"INDEKS source   : {os.path.basename(indeks_file)}")
    print(f"Bloomberg source: {os.path.basename(bb_file)}")

    # Read source data
    d1_df = pd.read_excel(d1_file, header=1)
    d2_df = pd.read_excel(d2_file, header=1)
    nav_28_df = pd.read_excel(nav_28_file, header=1)
    nav_27_df = pd.read_excel(nav_27_file, header=1)

    # Load fund_map for name overrides
    fund_map = load_fund_map()
    print(f"Fund map entries: {len(fund_map)}")

    # Get reference BRIMI names from compiled Excel (if available)
    if reference_path and os.path.exists(reference_path):
        xl_ref = pd.ExcelFile(reference_path)
        ref_d1 = pd.read_excel(xl_ref, "BRIMI D-1", header=1)
        ref_d2 = pd.read_excel(xl_ref, "BRIMI D-2", header=1)
        print(f"Reference BRIMI : {len(ref_d1)} funds from {os.path.basename(reference_path)}")
    else:
        # Fall back: use HistoricalNAV names directly
        # This works for funds where HistoricalNAV names match what process_brimi.py expects
        ref_d1 = nav_28_df[["Fund Name"]].rename(columns={"Fund Name": "Nama"})
        ref_d2 = nav_27_df[["Fund Name"]].rename(columns={"Fund Name": "Nama"})
        print(f"No reference    : using HistoricalNAV names directly ({len(ref_d1)} funds)")

    # Build BRIMI sheets
    brimi_d1_df = build_brimi_sheet(ref_d1, nav_28_df, fund_map, fallback_df=d1_df)
    brimi_d2_df = build_brimi_sheet(ref_d2, nav_27_df, fund_map, fallback_df=d2_df)
    print(f"BRIMI D-1       : {len(brimi_d1_df)} funds")
    print(f"BRIMI D-2       : {len(brimi_d2_df)} funds")

    # Extract date from header
    wb_hdr = openpyxl.load_workbook(d1_file, data_only=True)
    ws_hdr = wb_hdr["sheet 0"]
    nav_date = str(ws_hdr.cell(1, 1).value or "").replace("Data Per Tanggal : ", "").strip()
    wb_hdr.close()

    # Read INDEKS
    indeks_df = pd.read_excel(indeks_file, header=1)
    indeks_df = indeks_df.rename(
        columns={indeks_df.columns[0]: "Nama", indeks_df.columns[1]: "Nilai"}
    )

    # Copy BloombergIndex
    wb_bb_src = openpyxl.load_workbook(bb_file, data_only=True)

    # Write compiled output
    print(f"Writing         : {output_path}")
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    # D-1
    ws_d1 = wb_out.create_sheet("D-1")
    ws_d1.cell(1, 1, f"Data Per Tanggal : {nav_date}")
    for ci, col in enumerate(d1_df.columns, 1):
        ws_d1.cell(2, ci, col)
    for ri, row in enumerate(d1_df.values, 3):
        for ci, val in enumerate(row, 1):
            ws_d1.cell(ri, ci, val)
    print(f"D-1             : {len(d1_df)} funds")

    # D-2
    ws_d2 = wb_out.create_sheet("D-2")
    ws_d2.cell(1, 1, "Data Per Tanggal : 24-Apr")
    for ci, col in enumerate(d2_df.columns, 1):
        ws_d2.cell(2, ci, col)
    for ri, row in enumerate(d2_df.values, 3):
        for ci, val in enumerate(row, 1):
            ws_d2.cell(ri, ci, val)
    print(f"D-2             : {len(d2_df)} funds")

    # BRIMI D-1 — header at row 2 (header=1 in pandas)
    ws_bd1 = wb_out.create_sheet("BRIMI D-1")
    for ci in range(1, 7):
        ws_bd1.cell(1, ci, "")
    for ci, col in enumerate(brimi_d1_df.columns, 1):
        ws_bd1.cell(2, ci, col)
    for ri, row in enumerate(brimi_d1_df.values, 3):
        for ci, val in enumerate(row, 1):
            ws_bd1.cell(ri, ci, val)
    print(f"BRIMI D-1       : {len(brimi_d1_df)} funds")

    # BRIMI D-2
    ws_bd2 = wb_out.create_sheet("BRIMI D-2")
    for ci in range(1, 7):
        ws_bd2.cell(1, ci, "")
    for ci, col in enumerate(brimi_d2_df.columns, 1):
        ws_bd2.cell(2, ci, col)
    for ri, row in enumerate(brimi_d2_df.values, 3):
        for ci, val in enumerate(row, 1):
            ws_bd2.cell(ri, ci, val)
    print(f"BRIMI D-2       : {len(brimi_d2_df)} funds")

    # INDEKS
    ws_in = wb_out.create_sheet("INDEKS")
    ws_in.cell(1, 1, f"Data Per {nav_date}")
    for ci, col in enumerate(indeks_df.columns, 1):
        ws_in.cell(2, ci, col)
    for ri, row in enumerate(indeks_df.values, 3):
        for ci, val in enumerate(row, 1):
            ws_in.cell(ri, ci, val)
    print(f"INDEKS          : {len(indeks_df)} entries")

    # BloombergIndex
    ws_bb_src = wb_bb_src["Sheet1"]
    ws_bb = wb_out.create_sheet("BloombergIndex")
    for row_idx in range(1, ws_bb_src.max_row + 1):
        for col_idx in range(1, ws_bb_src.max_column + 1):
            src_cell = ws_bb_src.cell(row_idx, col_idx)
            ws_bb.cell(row_idx, col_idx).value = src_cell.value

    wb_out.save(output_path)
    wb_bb_src.close()
    print("Done.")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    ref = sys.argv[2] if len(sys.argv) > 2 else None
    build_compiled(sys.argv[1], ref)
